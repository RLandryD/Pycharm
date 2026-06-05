"""
reporter/report_generator.py
Produces two outputs from a list of MigrationAssessments:
  1. output/gap_analysis.xlsx  — detailed Excel with per-target destination sheets
  2. output/migration_report.md — Markdown stakeholder summary
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from analyzer.complexity_analyzer import MigrationAssessment

logger = logging.getLogger(__name__)

FILL_COLOURS = {
    "LOW":    "C6EFCE",
    "MEDIUM": "FFEB9C",
    "HIGH":   "FFC7CE",
}

BASE_HEADERS = [
    "Interface Name", "Namespace", "Sender System", "Sender Adapter",
    "Receiver System", "Receiver Adapter", "Has BPM", "Has Multi-Mapping",
    "Mapping Program", "Complexity", "Score", "Effort (days)",
    "Recommended Pattern", "Migration Notes",
]

TARGET_HEADERS = [
    "Interface Name", "Sender (Original→CPI)", "Receiver (Original→CPI)",
    "Supported?", "Cloud Connector?", "Effort Multiplier",
    "Hub Matches", "Compatibility Warnings", "Migration Hints",
]


class ReportGenerator:

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    def generate_excel(
        self,
        assessments: list[MigrationAssessment],
        resolutions: Optional[dict] = None,
        target_ids: Optional[list[str]] = None,
    ) -> Path:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl required: pip install openpyxl")

        wb = openpyxl.Workbook()

        # Summary sheet (index 0)
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary, assessments, target_ids or [])

        # Base inventory sheet
        ws_inv = wb.create_sheet("Interface Inventory")
        self._write_inventory_sheet(ws_inv, assessments)

        # One sheet per destination target
        if resolutions and target_ids:
            for tid in target_ids:
                try:
                    from destinations.registry import DESTINATION_REGISTRY
                    import re as _re
                    label = DESTINATION_REGISTRY.get(tid)
                    raw = label.label if label else tid
                    sheet_name = _re.sub(r'[\/\[\]:*?]', '-', raw)[:31]
                except Exception:
                    sheet_name = tid[:31]
                ws_target = wb.create_sheet(sheet_name)
                self._write_target_sheet(ws_target, assessments, resolutions, tid)

        out_path = self.output_dir / "gap_analysis.xlsx"
        wb.save(out_path)
        logger.info("Excel report saved → %s", out_path)
        return out_path

    def _write_inventory_sheet(self, ws, assessments: list[MigrationAssessment]):
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, header in enumerate(BASE_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

        for row_num, a in enumerate(assessments, start=2):
            iface = a.interface
            row_data = [
                iface.name, iface.namespace, iface.sender_system,
                iface.sender_adapter, iface.receiver_system, iface.receiver_adapter,
                "Yes" if iface.has_bpm else "No",
                "Yes" if iface.has_multi_mapping else "No",
                iface.mapping_program or "",
                a.complexity, a.score, a.effort_days,
                a.recommended_pattern, "\n".join(a.notes),
            ]
            fill_hex = FILL_COLOURS.get(a.complexity, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_hex)
            complexity_col = BASE_HEADERS.index("Complexity") + 1

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(col == len(BASE_HEADERS)))
                if col == complexity_col:
                    cell.fill = row_fill
                    cell.font = Font(bold=True)

        col_widths = [40, 30, 18, 14, 18, 14, 8, 14, 30, 10, 7, 12, 40, 55]
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_target_sheet(self, ws, assessments, resolutions, tid):
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return

        header_fill = PatternFill("solid", fgColor="0070B8")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, header in enumerate(TARGET_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

        warn_fill  = PatternFill("solid", fgColor="FFC7CE")
        ok_fill    = PatternFill("solid", fgColor="C6EFCE")

        for row_num, a in enumerate(assessments, start=2):
            iface = a.interface
            iface_res = resolutions.get(iface.name, {})
            resolved  = iface_res.get(tid)

            if resolved:
                sr = resolved.sender_recommendation
                rr = resolved.receiver_recommendation
                supported = "✓" if (sr.is_supported and rr.is_supported) else "⚠ No"
                needs_cc  = "Yes" if (sr.requires_cloud_connector or rr.requires_cloud_connector) else "No"
                hub_text  = "\n".join(
                    f"[{m.artifact_type}] {m.title}" for m in resolved.hub_matches
                ) if resolved.hub_matches else "None found"
                warnings  = "\n".join(resolved.compatibility_warnings) or "None"
                hints     = "\n".join(resolved.migration_hints[:4]) if resolved.migration_hints else ""
                multiplier = f"×{resolved.effort_multiplier:.1f}"
                sender_str   = f"{sr.original_adapter} → {sr.recommended_adapter}"
                receiver_str = f"{rr.original_adapter} → {rr.recommended_adapter}"
            else:
                sender_str = receiver_str = supported = needs_cc = hub_text = warnings = hints = multiplier = "N/A"

            row_data = [
                iface.name, sender_str, receiver_str,
                supported, needs_cc, multiplier,
                hub_text, warnings, hints,
            ]
            is_ok = resolved and not resolved.compatibility_warnings if resolved else False

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col == 4:  # Supported column
                    cell.fill = ok_fill if is_ok else warn_fill

        col_widths = [38, 28, 28, 12, 16, 14, 45, 50, 55]
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_summary_sheet(self, ws, assessments, target_ids: list[str]):
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return

        total = len(assessments)
        low    = sum(1 for a in assessments if a.complexity == "LOW")
        medium = sum(1 for a in assessments if a.complexity == "MEDIUM")
        high   = sum(1 for a in assessments if a.complexity == "HIGH")
        total_days = sum(a.effort_days for a in assessments)

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 22

        ws["A1"] = "PI/PO → CPI Migration Assessment"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(italic=True, color="808080")

        rows = [
            ("", ""),
            ("Metric", "Value"),
            ("Total interfaces", total),
            ("Low complexity",    low),
            ("Medium complexity", medium),
            ("High complexity",   high),
            ("Total effort estimate", f"{total_days:.1f} days"),
            ("", ""),
            ("Destination targets", ", ".join(target_ids) if target_ids else "None"),
        ]
        for r, (label, value) in enumerate(rows, start=3):
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=value)
            if label == "Metric":
                for c in [1, 2]:
                    ws.cell(row=r, column=c).font = Font(bold=True)

        from openpyxl.styles import PatternFill as PF
        for row_offset, colour in enumerate(["C6EFCE", "FFEB9C", "FFC7CE"], start=1):
            ws.cell(row=5 + row_offset, column=2).fill = PF("solid", fgColor=colour)

    # ------------------------------------------------------------------
    # Markdown
    # ------------------------------------------------------------------

    def generate_markdown(
        self,
        assessments: list[MigrationAssessment],
        resolutions: Optional[dict] = None,
        target_ids: Optional[list[str]] = None,
    ) -> Path:
        total = len(assessments)
        low    = sum(1 for a in assessments if a.complexity == "LOW")
        medium = sum(1 for a in assessments if a.complexity == "MEDIUM")
        high   = sum(1 for a in assessments if a.complexity == "HIGH")
        total_days = sum(a.effort_days for a in assessments)

        lines = [
            "# PI/PO → CPI Migration Assessment Report",
            f"\n_Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}_\n",
            "## Executive Summary\n",
            "| Metric | Value |",
            "|--------|-------|",
            f"| Total interfaces | **{total}** |",
            f"| 🟢 Low complexity | {low} |",
            f"| 🟡 Medium complexity | {medium} |",
            f"| 🔴 High complexity | {high} |",
            f"| **Total effort estimate** | **{total_days:.1f} days** |",
        ]

        if target_ids:
            lines += [
                f"| Destination targets | {', '.join(f'`{t}`' for t in target_ids)} |",
            ]

        lines += ["\n---\n", "## Interface Inventory\n",
                  "| Interface | Sender | Receiver | Complexity | Effort | Pattern |",
                  "|-----------|--------|----------|------------|--------|---------|"]

        emoji = {"LOW": "🟢", "MEDIUM": "🟡", "HIGH": "🔴"}
        for a in sorted(assessments, key=lambda x: x.score, reverse=True):
            iface = a.interface
            lines.append(
                f"| {iface.name} "
                f"| {iface.sender_system} ({iface.sender_adapter}) "
                f"| {iface.receiver_system} ({iface.receiver_adapter}) "
                f"| {emoji.get(a.complexity,'')} {a.complexity} "
                f"| {a.effort_days}d "
                f"| {a.recommended_pattern} |"
            )

        # Per-target destination sections
        if resolutions and target_ids:
            for tid in target_ids:
                try:
                    from destinations.registry import DESTINATION_REGISTRY
                    label = getattr(DESTINATION_REGISTRY.get(tid), 'label', tid)
                except Exception:
                    label = tid

                lines += [f"\n---\n", f"## Destination: {label}\n",
                          "| Interface | Sender→CPI | Receiver→CPI | Warnings | Hub Matches |",
                          "|-----------|-----------|--------------|----------|-------------|"]

                for a in sorted(assessments, key=lambda x: x.score, reverse=True):
                    iface = a.interface
                    resolved = resolutions.get(iface.name, {}).get(tid)
                    if not resolved:
                        continue
                    sr = resolved.sender_recommendation
                    rr = resolved.receiver_recommendation
                    warn_str  = f"⚠ {len(resolved.compatibility_warnings)}" if resolved.compatibility_warnings else "✓"
                    hub_str   = str(len(resolved.hub_matches)) if resolved.hub_matches else "0"
                    lines.append(
                        f"| {iface.name} "
                        f"| {sr.original_adapter}→{sr.recommended_adapter} "
                        f"| {rr.original_adapter}→{rr.recommended_adapter} "
                        f"| {warn_str} "
                        f"| {hub_str} |"
                    )

                # Hub matches detail for this target
                all_matches = [
                    (a.interface.name, m)
                    for a in assessments
                    for m in resolutions.get(a.interface.name, {}).get(tid, type('', (), {'hub_matches': []})()).hub_matches
                ]
                if all_matches:
                    lines += [f"\n### Pre-built Hub content for {label}\n"]
                    for iface_name, m in all_matches[:10]:
                        lines.append(f"- **{iface_name}** — [{m.title}]({m.url}) `{m.artifact_type}`")

        # HIGH complexity detail
        high_items = [a for a in assessments if a.complexity == "HIGH"]
        if high_items:
            lines += ["\n---\n", "## High Complexity — Action Required\n"]
            for a in high_items:
                lines += [
                    f"### {a.interface.name}",
                    f"- **Score:** {a.score} | **Effort:** {a.effort_days} days",
                    f"- **Pattern:** {a.recommended_pattern}",
                ]
                if a.notes:
                    lines.append("- **Notes:**")
                    for note in a.notes:
                        lines.append(f"  - {note}")
                lines.append("")

        out_path = self.output_dir / "migration_report.md"
        out_path.write_text("\n".join(lines), encoding="utf-8")
        logger.info("Markdown report saved → %s", out_path)
        return out_path
