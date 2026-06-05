"""
reporter/intervention_estimator.py

Calculates exactly what the tool has automated vs what requires
human intervention for each interface. Produces:

  1. Per-interface intervention card (automation %, hours, task list)
  2. Project summary (total hours, missing info, blocked interfaces)
  3. Client info request sheet (what client must provide)
  4. Internal effort breakdown (your hours + cost + margin)

Reads from existing analysis objects — no new data collection needed.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

from reporter.effort_model import (
    EffortBreakdown, build_effort, snap_multiplier,
    MULTIPLIER_DEFAULT, HYPERCARE_DEFAULT_HOURS,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Time estimates per task type (hours)
# ---------------------------------------------------------------------------

TASK_HOURS = {
    # Automated (tool does it) — shown as ✓
    "iflow_scaffold":          (0.0,  "iFlow XML structure generated"),
    "package_naming":          (0.0,  "Package and iFlow naming applied"),
    "exception_handler":       (0.0,  "Exception handler Groovy script generated"),
    "tdd_doc":                 (0.0,  "TDD documentation generated"),
    "preflight_checklist":     (0.0,  "Pre-flight checklist generated"),
    "test_harness":            (0.0,  "Test harness + mock payload generated"),
    "security_inventory":      (0.0,  "Security inventory entry created"),
    "clean_core_check":        (0.0,  "Clean Core compliance scored"),

    # Manual — shown as ⏱
    "message_mapping":         (1.5,  "Implement message mapping in CPI Message Mapping editor"),
    "groovy_mapping_simple":   (1.0,  "Complete Groovy transformation script (simple fields)"),
    "groovy_mapping_complex":  (3.0,  "Complete Groovy transformation script (complex logic)"),
    "credential_create":       (0.3,  "Create secure parameter / credential entry in CPI"),
    "oauth_setup":             (0.5,  "Configure OAuth2 client credential in CPI"),
    "certificate_import":      (0.5,  "Import certificate into CPI keystore"),
    "idoc_we20":               (0.5,  "Configure WE20 partner profile in SAP system"),
    "idoc_we21":               (0.3,  "Configure WE21 port in SAP system"),
    "rfc_sm59":                (0.3,  "Create SM59 RFC destination"),
    "scc_system_mapping":      (0.5,  "Add system mapping in Cloud Connector"),
    "jdbc_datasource":         (0.5,  "Configure JDBC data source in CPI"),
    "jdbc_firewall":           (0.5,  "Open firewall from CPI to database (client IT)"),
    "sftp_key_exchange":       (0.5,  "Exchange SFTP keys and test connectivity"),
    "as2_partner_setup":       (2.0,  "Configure AS2 partner profile and certificate exchange"),
    "s4_comm_arrangement":     (1.0,  "Create Communication Arrangement in S/4HANA"),
    "business_test":           (1.0,  "End-to-end business test with real payload"),
    "test_sign_off":           (0.5,  "Business user test sign-off"),
    "deploy_transport":        (0.5,  "Deploy iFlow and transport to QA"),
    "bpm_redesign":            (4.0,  "Redesign BPM process as CPI iFlow steps"),
    "multi_mapping_redesign":  (1.5,  "Implement Multicast/Splitter for multi-mapping"),
    "java_rewrite":            (3.0,  "Rewrite Java mapping as native Groovy"),
    "eoio_datastore":          (1.0,  "Configure EOIO DataStore pattern and test ordering"),
    "undocumented_research":   (4.0,  "Reverse-engineer undocumented interface logic"),
    "find_odata_equivalent":   (1.5,  "Find OData/SOAP API equivalent for RFC/BAPI on Hub"),
    "namespace_config":        (0.3,  "Configure namespace mapping"),
}

# What triggers each manual task
TASK_TRIGGERS = {
    "message_mapping":        lambda iface, cfg, vr, cc: bool(iface.mapping_program),
    "groovy_mapping_simple":  lambda iface, cfg, vr, cc: (
        bool(iface.mapping_program) and
        iface.sender_adapter not in ("IDoc",) and
        not iface.has_bpm
    ),
    "groovy_mapping_complex": lambda iface, cfg, vr, cc: (
        bool(iface.mapping_program) and iface.has_bpm
    ),
    "credential_create":      lambda iface, cfg, vr, cc: (
        cfg and cfg.sender_auth.method == "Basic" and
        not cfg.sender_auth.credential_name
    ),
    "oauth_setup":            lambda iface, cfg, vr, cc: (
        cfg and cfg.receiver_auth.method == "OAuth2 Client Credentials"
    ),
    "certificate_import":     lambda iface, cfg, vr, cc: (
        cfg and cfg.receiver_auth.method == "Certificate"
    ),
    "idoc_we20":              lambda iface, cfg, vr, cc: (
        iface.sender_adapter == "IDoc" or iface.receiver_adapter == "IDoc"
    ),
    "idoc_we21":              lambda iface, cfg, vr, cc: (
        iface.sender_adapter == "IDoc" or iface.receiver_adapter == "IDoc"
    ),
    "rfc_sm59":               lambda iface, cfg, vr, cc: (
        iface.sender_adapter == "RFC" or iface.receiver_adapter == "RFC"
    ),
    "scc_system_mapping":     lambda iface, cfg, vr, cc: (
        iface.sender_adapter in ("RFC","JDBC","IDoc","File","FTP","SFTP") or
        iface.receiver_adapter in ("RFC","JDBC")
    ),
    "jdbc_datasource":        lambda iface, cfg, vr, cc: (
        iface.sender_adapter == "JDBC" or iface.receiver_adapter == "JDBC"
    ),
    "jdbc_firewall":          lambda iface, cfg, vr, cc: (
        iface.sender_adapter == "JDBC" or iface.receiver_adapter == "JDBC"
    ),
    "sftp_key_exchange":      lambda iface, cfg, vr, cc: (
        iface.sender_adapter in ("SFTP","FTP") or
        iface.receiver_adapter in ("SFTP","FTP")
    ),
    "as2_partner_setup":      lambda iface, cfg, vr, cc: (
        iface.sender_adapter in ("AS2","AS4") or
        iface.receiver_adapter in ("AS2","AS4")
    ),
    "s4_comm_arrangement":    lambda iface, cfg, vr, cc: True,
    "business_test":          lambda iface, cfg, vr, cc: True,
    "test_sign_off":          lambda iface, cfg, vr, cc: True,
    "deploy_transport":       lambda iface, cfg, vr, cc: True,
    "bpm_redesign":           lambda iface, cfg, vr, cc: iface.has_bpm,
    "multi_mapping_redesign": lambda iface, cfg, vr, cc: iface.has_multi_mapping,
    "java_rewrite":           lambda iface, cfg, vr, cc: (
        bool(iface.mapping_program) and
        any(kw in iface.mapping_program.lower()
            for kw in ("java","jar","binary","pdf","excel"))
    ),
    "eoio_datastore":         lambda iface, cfg, vr, cc: (
        iface.sender_adapter in ("JMS","XI") or
        iface.receiver_adapter in ("JMS","XI")
    ),
    "undocumented_research":  lambda iface, cfg, vr, cc: (
        not iface.description and not iface.mapping_program and
        not iface.message_interface
    ),
    "find_odata_equivalent":  lambda iface, cfg, vr, cc: (
        iface.sender_adapter == "RFC" or iface.receiver_adapter == "RFC"
    ),
    "namespace_config":       lambda iface, cfg, vr, cc: bool(iface.namespace),
}

# Responsible party per task
TASK_RESPONSIBLE = {
    "message_mapping":        "You",
    "groovy_mapping_simple":  "You",
    "groovy_mapping_complex": "You",
    "credential_create":      "You",
    "oauth_setup":            "You",
    "certificate_import":     "You + Client Security",
    "idoc_we20":              "Client Basis",
    "idoc_we21":              "Client Basis",
    "rfc_sm59":               "Client Basis",
    "scc_system_mapping":     "Client Basis",
    "jdbc_datasource":        "You",
    "jdbc_firewall":          "Client IT/Security",
    "sftp_key_exchange":      "You + Client",
    "as2_partner_setup":      "You + External Partner",
    "s4_comm_arrangement":    "Client Basis",
    "business_test":          "You + Client Business",
    "test_sign_off":          "Client Business",
    "deploy_transport":       "You",
    "bpm_redesign":           "You (senior work)",
    "multi_mapping_redesign": "You",
    "java_rewrite":           "You",
    "eoio_datastore":         "You",
    "undocumented_research":  "You + Client SME",
    "find_odata_equivalent":  "You",
    "namespace_config":       "You",
}


# ---------------------------------------------------------------------------
# Output models
# ---------------------------------------------------------------------------

@dataclass
class InterventionTask:
    task_id: str
    description: str
    hours: float
    responsible: str
    is_automated: bool
    is_blocking: bool = False


@dataclass
class MissingInfo:
    field: str
    description: str
    blocking: bool
    client_action: str


@dataclass
class InterfaceIntervention:
    interface_name: str
    automation_pct: float
    total_manual_hours: float
    your_hours: float
    client_hours: float
    automated_tasks: list[InterventionTask]
    manual_tasks: list[InterventionTask]
    missing_info: list[MissingInfo]
    tier: str = "GUIDED"           # AUTO / GUIDED / SPECIALIST
    ready_to_start: bool = True
    effort: Optional[EffortBreakdown] = None   # base+gaps+multiplier+hypercare


@dataclass
class ProjectIntervention:
    total_interfaces: int
    total_manual_hours: float
    your_hours: float
    client_hours: float
    avg_automation_pct: float
    ready_to_start: int
    blocked_on_client: int
    specialist_count: int
    missing_info_total: int
    blocking_missing: int
    interfaces: list[InterfaceIntervention] = field(default_factory=list)
    effort: Optional[EffortBreakdown] = None   # project-level rolled-up effort


# ---------------------------------------------------------------------------
# Estimator
# ---------------------------------------------------------------------------

class InterventionEstimator:

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir) / "intervention"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def estimate(
        self,
        assessment,
        cfg=None,
        verification_report=None,
        clean_core_report=None,
        ceiling=None,
        engine_result=None,
        multiplier: float = MULTIPLIER_DEFAULT,
        mode: str = "",
        hypercare_enabled: bool = False,
        hypercare_hours: float = HYPERCARE_DEFAULT_HOURS,
    ) -> InterfaceIntervention:
        iface = assessment.interface
        vr    = verification_report
        cc    = clean_core_report

        # Automated tasks (always done by tool)
        automated = [
            InterventionTask(
                task_id=k, description=desc,
                hours=0.0, responsible="Tool",
                is_automated=True,
            )
            for k, (_, desc) in [
                ("iflow_scaffold",      TASK_HOURS["iflow_scaffold"]),
                ("exception_handler",   TASK_HOURS["exception_handler"]),
                ("tdd_doc",             TASK_HOURS["tdd_doc"]),
                ("preflight_checklist", TASK_HOURS["preflight_checklist"]),
                ("test_harness",        TASK_HOURS["test_harness"]),
                ("security_inventory",  TASK_HOURS["security_inventory"]),
                ("clean_core_check",    TASK_HOURS["clean_core_check"]),
            ]
        ]
        if iface.mapping_program:
            automated.append(InterventionTask(
                task_id="groovy_stub",
                description="Groovy script stub generated (logic needs completion)",
                hours=0.0, responsible="Tool", is_automated=True,
            ))

        # Manual tasks — evaluated per trigger
        manual = []
        for task_id, trigger_fn in TASK_TRIGGERS.items():
            try:
                if trigger_fn(iface, cfg, vr, cc):
                    hours, desc = TASK_HOURS[task_id]
                    responsible = TASK_RESPONSIBLE[task_id]
                    manual.append(InterventionTask(
                        task_id=task_id,
                        description=desc,
                        hours=hours,
                        responsible=responsible,
                        is_automated=False,
                        is_blocking=task_id in (
                            "bpm_redesign", "java_rewrite",
                            "undocumented_research", "as2_partner_setup",
                        ),
                    ))
            except Exception:
                pass

        # Missing info items
        missing = self._detect_missing_info(iface, cfg, vr)

        # Hours breakdown
        total_manual    = sum(t.hours for t in manual)
        your_hours      = sum(
            t.hours for t in manual
            if t.responsible in ("You", "You + Client Security",
                                  "You + Client", "You + External Partner",
                                  "You + Client Business")
        )
        client_hours    = total_manual - your_hours

        # Automation %
        total_tasks     = len(automated) + len(manual)
        automation_pct  = (len(automated) / total_tasks * 100) if total_tasks else 0

        # Tier from ceiling if available
        tier = "GUIDED"
        if ceiling:
            tier = ceiling.tier
        elif iface.has_bpm:
            tier = "SPECIALIST"
        elif assessment.complexity == "HIGH":
            tier = "GUIDED"
        else:
            tier = "AUTO" if automation_pct >= 80 else "GUIDED"

        # Ready to start?
        blocking_missing = [m for m in missing if m.blocking]
        ready            = len(blocking_missing) == 0 and tier != "SPECIALIST"

        # --- Effort breakdown: engine MA base (before automation) + gap hours ---
        # BASE = engine Mode-1 assessment of this interface (weight->size->
        # category->hours). NEVER the finished bundle. If no engine_result was
        # supplied (legacy callers), base is 0 and the breakdown still reflects
        # the itemized gap hours so nothing is silently lost.
        if engine_result is not None:
            base_lo = float(getattr(engine_result, "effort_hours_low", 0.0))
            base_hi = float(getattr(engine_result, "effort_hours_high", 0.0))
        else:
            base_lo = base_hi = 0.0
        # GAP HOURS = your-side itemized manual tasks (the research/testing/
        # redesign work), already computed above as your_hours.
        effort = build_effort(
            base_hours_low=base_lo,
            base_hours_high=base_hi,
            gap_hours=your_hours,
            client_hours=client_hours,
            multiplier=multiplier,
            mode=mode,
            hypercare_enabled=hypercare_enabled,
            hypercare_hours=hypercare_hours,
        )

        return InterfaceIntervention(
            interface_name=iface.name,
            automation_pct=automation_pct,
            total_manual_hours=total_manual,
            your_hours=your_hours,
            client_hours=client_hours,
            automated_tasks=automated,
            manual_tasks=manual,
            missing_info=missing,
            tier=tier,
            ready_to_start=ready,
            effort=effort,
        )

    def estimate_all(
        self,
        assessments: list,
        configs: dict = None,
        verification_reports: dict = None,
        clean_core_reports: dict = None,
        ceilings: dict = None,
        engine_results: dict = None,
        multiplier: float = MULTIPLIER_DEFAULT,
        mode: str = "",
        hypercare_enabled: bool = False,
        hypercare_hours: float = HYPERCARE_DEFAULT_HOURS,
    ) -> ProjectIntervention:
        configs              = configs or {}
        verification_reports = verification_reports or {}
        clean_core_reports   = clean_core_reports or {}
        ceilings             = ceilings or {}
        engine_results       = engine_results or {}

        interventions = []
        for a in assessments:
            name = a.interface.name
            iv   = self.estimate(
                a,
                cfg=configs.get(name),
                verification_report=verification_reports.get(name),
                clean_core_report=clean_core_reports.get(name),
                ceiling=ceilings.get(name),
                engine_result=engine_results.get(name),
                multiplier=multiplier,
                mode=mode,
                hypercare_enabled=False,   # hypercare is added once, project-level
                hypercare_hours=0.0,
            )
            interventions.append(iv)

        total_manual   = sum(iv.total_manual_hours for iv in interventions)
        your_hours     = sum(iv.your_hours for iv in interventions)
        client_hours   = sum(iv.client_hours for iv in interventions)
        avg_auto       = (sum(iv.automation_pct for iv in interventions) /
                          len(interventions)) if interventions else 0
        ready          = sum(1 for iv in interventions if iv.ready_to_start)
        blocked        = sum(1 for iv in interventions if not iv.ready_to_start)
        specialist     = sum(1 for iv in interventions if iv.tier == "SPECIALIST")
        missing_total  = sum(len(iv.missing_info) for iv in interventions)
        blocking_miss  = sum(
            sum(1 for m in iv.missing_info if m.blocking)
            for iv in interventions
        )

        # --- Project-level effort roll-up ---
        # Sum the per-interface base ranges and gap hours, apply the multiplier
        # once, and add hypercare once (flat, project-level, optional).
        proj_base_lo = sum(
            iv.effort.base_hours_low for iv in interventions if iv.effort)
        proj_base_hi = sum(
            iv.effort.base_hours_high for iv in interventions if iv.effort)
        proj_gap = sum(iv.effort.gap_hours for iv in interventions if iv.effort)
        proj_client = sum(
            iv.effort.client_hours for iv in interventions if iv.effort)
        project_effort = build_effort(
            base_hours_low=proj_base_lo,
            base_hours_high=proj_base_hi,
            gap_hours=proj_gap,
            client_hours=proj_client,
            multiplier=multiplier,
            mode=mode,
            hypercare_enabled=hypercare_enabled,
            hypercare_hours=hypercare_hours,
        )

        return ProjectIntervention(
            total_interfaces=len(interventions),
            total_manual_hours=total_manual,
            your_hours=your_hours,
            client_hours=client_hours,
            avg_automation_pct=avg_auto,
            ready_to_start=ready,
            blocked_on_client=blocked,
            specialist_count=specialist,
            missing_info_total=missing_total,
            blocking_missing=blocking_miss,
            interfaces=interventions,
            effort=project_effort,
        )

    def _detect_missing_info(self, iface, cfg, vr) -> list[MissingInfo]:
        missing = []

        if cfg:
            if not cfg.receiver_connectivity.address:
                missing.append(MissingInfo(
                    field="Receiver address",
                    description=f"Target system URL/host for {iface.receiver_system} not set",
                    blocking=True,
                    client_action=f"Provide {iface.receiver_system} API endpoint URL",
                ))
            if not cfg.receiver_auth.credential_name and cfg.receiver_auth.method != "None":
                missing.append(MissingInfo(
                    field="Receiver credential name",
                    description="Credential store alias not specified",
                    blocking=False,
                    client_action="Provide service key / credential details for target system",
                ))
            if cfg.receiver_auth.method == "OAuth2 Client Credentials" and not cfg.receiver_auth.token_url:
                missing.append(MissingInfo(
                    field="OAuth2 token URL",
                    description="Token URL missing for OAuth2 auth",
                    blocking=True,
                    client_action="Provide OAuth2 token URL from target system service key",
                ))

        if not iface.mapping_program and iface.sender_adapter != iface.receiver_adapter:
            missing.append(MissingInfo(
                field="Mapping program",
                description="No message mapping specified — may need transformation",
                blocking=False,
                client_action="Confirm if field mapping is required or pass-through",
            ))

        if iface.sender_adapter in ("IDoc","RFC") and not iface.sender_system:
            missing.append(MissingInfo(
                field="Sender system",
                description="Source SAP system host not identified",
                blocking=True,
                client_action="Provide source SAP system hostname and system number",
            ))

        if vr:
            for gap in vr.gaps:
                if gap.severity == "BLOCKING" and "address" not in gap.description.lower():
                    missing.append(MissingInfo(
                        field=gap.category,
                        description=gap.description,
                        blocking=True,
                        client_action=gap.suggested_fix,
                    ))

        return missing

    # ── Excel output ──────────────────────────────────────────────────

    def generate_excel(
        self,
        project: ProjectIntervention,
        project_name: str = "CPI Migration",
        day_rate_usd: float = 100.0,
        price_low: float = 1250.0,
        price_medium: float = 2500.0,
        price_high: float = 5500.0,
    ) -> Path:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl required")

        wb   = openpyxl.Workbook()
        thin = Side(style="thin")
        bdr  = openpyxl.styles.Border(
            left=thin, right=thin, top=thin, bottom=thin)

        def hdr(ws, row, col, val, colour="1F4E79"):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=colour)
            c.font = Font(bold=True, color="FFFFFF")
            c.border = bdr
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        def dat(ws, row, col, val, bold=False, colour=None, wrap=False):
            c = ws.cell(row=row, column=col, value=val)
            c.border = bdr
            c.alignment = Alignment(vertical="top", wrap_text=wrap)
            if bold: c.font = Font(bold=True)
            if colour: c.fill = PatternFill("solid", fgColor=colour)

        # ── Sheet 1: Project Summary ──────────────────────────────────
        ws1 = wb.active
        ws1.title = "Project Summary"
        ws1["A1"] = f"Human Intervention Estimate — {project_name}"
        ws1["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws1["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
        ws1["A2"].font = Font(italic=True, color="808080")

        summary_rows = [
            ("Total interfaces",           project.total_interfaces),
            ("Average automation",         f"{project.avg_automation_pct:.0f}%"),
            ("Total manual hours",         f"{project.total_manual_hours:.1f} hrs"),
            ("  → Your hours",             f"{project.your_hours:.1f} hrs"),
            ("  → Client hours",           f"{project.client_hours:.1f} hrs"),
            ("Ready to start now",         project.ready_to_start),
            ("Blocked on client info",     project.blocked_on_client),
            ("Needs specialist",           project.specialist_count),
            ("Missing info fields total",  project.missing_info_total),
            ("  → Blocking fields",        project.blocking_missing),
        ]
        for r, (label, value) in enumerate(summary_rows, 4):
            ws1.cell(row=r, column=1, value=label)
            ws1.cell(row=r, column=2, value=value)
            if "→" in label:
                ws1.cell(row=r, column=1).font = Font(color="595959")

        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 18

        # ── Sheet 2: Interface Breakdown (your view) ──────────────────
        ws2 = wb.create_sheet("Interface Breakdown")
        headers2 = ["Interface", "Tier", "Automation %", "Your Hours",
                    "Client Hours", "Total Manual", "Missing Info",
                    "Blocking?", "Ready?", "Manual Tasks Summary"]
        for col, h in enumerate(headers2, 1):
            hdr(ws2, 1, col, h)

        tier_colours = {
            "AUTO":       "C6EFCE",
            "GUIDED":     "FFEB9C",
            "SPECIALIST": "FFC7CE",
        }
        for row, iv in enumerate(project.interfaces, 2):
            colour = tier_colours.get(iv.tier, "FFFFFF")
            tasks_summary = "; ".join(
                f"{t.description[:30]}" for t in iv.manual_tasks[:3]
            )
            row_data = [
                iv.interface_name, iv.tier,
                f"{iv.automation_pct:.0f}%",
                f"{iv.your_hours:.1f}",
                f"{iv.client_hours:.1f}",
                f"{iv.total_manual_hours:.1f}",
                len(iv.missing_info),
                "YES" if any(m.blocking for m in iv.missing_info) else "",
                "✓" if iv.ready_to_start else "⚠ Blocked",
                tasks_summary,
            ]
            for col, val in enumerate(row_data, 1):
                dat(ws2, row, col, val,
                    colour=colour if col in (1, 2) else None,
                    wrap=(col == 10))

        for col, width in enumerate([38,12,14,12,14,14,13,10,12,55], 1):
            ws2.column_dimensions[get_column_letter(col)].width = width
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"

        # ── Sheet 3: Client Info Request ──────────────────────────────
        ws3 = wb.create_sheet("Client Info Request")
        ws3["A1"] = "Information Required from Client"
        ws3["A1"].font = Font(bold=True, size=13, color="1F4E79")
        ws3["A2"] = ("Please provide the following information before migration "
                     "can begin. Items marked BLOCKING must be resolved first.")
        ws3["A2"].font = Font(italic=True, color="808080")

        headers3 = ["Interface", "Missing Field", "Description",
                    "Client Action Required", "Blocking?", "Provided? (✓)"]
        for col, h in enumerate(headers3, 1):
            hdr(ws3, 4, col, h, "C55A11")

        row = 5
        for iv in project.interfaces:
            for m in iv.missing_info:
                colour = "FFC7CE" if m.blocking else "FFFADD"
                row_data = [
                    iv.interface_name, m.field,
                    m.description, m.client_action,
                    "BLOCKING" if m.blocking else "optional",
                    "",
                ]
                for col, val in enumerate(row_data, 1):
                    dat(ws3, row, col, val, colour=colour if col <= 2 else None,
                        wrap=(col in (3, 4)))
                row += 1

        for col, width in enumerate([35, 22, 45, 50, 12, 14], 1):
            ws3.column_dimensions[get_column_letter(col)].width = width
        ws3.freeze_panes = "A5"

        # ── Sheet 4: Your task detail ─────────────────────────────────
        ws4 = wb.create_sheet("Your Task Detail")
        ws4["A1"] = "Your Manual Tasks — Internal"
        ws4["A1"].font = Font(bold=True, size=13, color="1F4E79")

        headers4 = ["Interface", "Task", "Hours", "Responsible",
                    "Blocking?", "Notes"]
        for col, h in enumerate(headers4, 1):
            hdr(ws4, 3, col, h, "2E75B6")

        row = 4
        for iv in project.interfaces:
            for t in iv.automated_tasks:
                dat(ws4, row, 1, iv.interface_name)
                dat(ws4, row, 2, f"✓ {t.description}",
                    colour="C6EFCE")
                dat(ws4, row, 3, "Auto")
                dat(ws4, row, 4, "Tool")
                dat(ws4, row, 5, "")
                dat(ws4, row, 6, "")
                row += 1
            for t in iv.manual_tasks:
                dat(ws4, row, 1, iv.interface_name)
                dat(ws4, row, 2, f"⏱ {t.description}",
                    colour="FFC7CE" if t.is_blocking else "FFEB9C")
                dat(ws4, row, 3, t.hours)
                dat(ws4, row, 4, t.responsible)
                dat(ws4, row, 5, "BLOCKING" if t.is_blocking else "")
                dat(ws4, row, 6, "")
                row += 1

        for col, width in enumerate([35, 55, 8, 28, 10, 30], 1):
            ws4.column_dimensions[get_column_letter(col)].width = width
        ws4.freeze_panes = "A4"

        out = self.output_dir / "intervention_estimate.xlsx"
        wb.save(out)
        logger.info("Intervention estimate saved → %s", out)
        return out
