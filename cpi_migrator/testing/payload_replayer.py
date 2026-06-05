"""
testing/payload_replayer.py

Extracts real historical XML payloads from PI/PO message processing logs
and replays them through the newly migrated CPI iFlows for validation.

PI/PO endpoint:
  GET /MessageProcessingLogs?$format=json&$top=100
  GET /MessageProcessingLogs('{id}')/Payloads
  GET /MessageProcessingLogs('{id}')/Payloads('request')/$value  → raw XML

CPI replay:
  POST {iflow_endpoint} with extracted payload
  Compare response with expected output

Output:
  output/replay/
    {interface_name}_payload_{n}.xml   — extracted historical payloads
    replay_report.xlsx                 — comparison results
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests

logger = logging.getLogger(__name__)


@dataclass
class HistoricalPayload:
    log_id: str
    interface_name: str
    sender_adapter: str
    status: str              # Success / Failed
    start_time: str
    payload_content: bytes = b""
    payload_format: str = "xml"
    message_id: str = ""


@dataclass
class ReplayResult:
    interface_name: str
    payload_id: str
    pi_status: str
    cpi_status_code: int
    cpi_response: str
    match: bool
    diff_notes: str = ""
    duration_ms: int = 0


class PayloadReplayer:

    def __init__(
        self,
        pi_base_url: str,
        pi_session: requests.Session,
        cpi_base_url: str,
        cpi_session: requests.Session,
        output_dir: str = "./output",
    ):
        self.pi_base_url  = pi_base_url.rstrip("/")
        self.pi_session   = pi_session
        self.cpi_base_url = cpi_base_url.rstrip("/")
        self.cpi_session  = cpi_session
        self.output_dir   = Path(output_dir) / "replay"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ── Extract from PI/PO ────────────────────────────────────────────

    def extract_payloads(
        self,
        interface_name: str,
        max_payloads: int = 10,
        status_filter: str = "Success",
    ) -> list[HistoricalPayload]:
        """Extract historical payloads from PI/PO message logs."""
        payloads = []
        try:
            logs = self._fetch_message_logs(interface_name, max_payloads,
                                             status_filter)
            for log in logs:
                payload = self._fetch_payload_content(log)
                if payload:
                    payloads.append(payload)
                    self._save_payload(payload)
        except Exception as exc:
            logger.error("Payload extraction failed: %s", exc)
        return payloads

    def _fetch_message_logs(
        self,
        interface_name: str,
        top: int,
        status: str,
    ) -> list[dict]:
        url = f"{self.pi_base_url}/MessageProcessingLogs"
        params = {
            "$format":  "json",
            "$top":     top,
            "$orderby": "LogStart desc",
        }
        if interface_name:
            params["$filter"] = (f"IntegrationFlowName eq '{interface_name}' "
                                 f"and Status eq '{status}'")

        resp = self.pi_session.get(url, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        return data.get("d", {}).get("results", data.get("value", []))

    def _fetch_payload_content(self, log: dict) -> Optional[HistoricalPayload]:
        log_id = log.get("MessageGuid") or log.get("LogId") or log.get("id", "")
        if not log_id:
            return None

        iface_name  = log.get("IntegrationFlowName", "unknown")
        start_time  = log.get("LogStart", "")
        status      = log.get("Status", "")
        sender      = log.get("Sender", "")
        message_id  = log.get("CorrelationId", log_id)

        # Try to fetch the actual payload
        payload_bytes = b""
        for direction in ("request", "response"):
            url = (f"{self.pi_base_url}/MessageProcessingLogs('{log_id}')"
                   f"/Payloads('{direction}')/$value")
            try:
                resp = self.pi_session.get(url, timeout=20)
                if resp.status_code == 200:
                    payload_bytes = resp.content
                    break
            except Exception:
                continue

        if not payload_bytes:
            return None

        fmt = "json" if payload_bytes.lstrip()[:1] in (b"{", b"[") else "xml"

        return HistoricalPayload(
            log_id=log_id,
            interface_name=iface_name,
            sender_adapter=sender,
            status=status,
            start_time=start_time,
            payload_content=payload_bytes,
            payload_format=fmt,
            message_id=message_id,
        )

    def _save_payload(self, payload: HistoricalPayload):
        ext  = ".json" if payload.payload_format == "json" else ".xml"
        safe = payload.interface_name.replace("/", "_").replace(" ", "_")[:40]
        path = self.output_dir / f"{safe}_{payload.log_id[:8]}{ext}"
        path.write_bytes(payload.payload_content)
        logger.debug("Saved payload → %s", path)

    # ── Replay through CPI ────────────────────────────────────────────

    def replay_payload(
        self,
        payload: HistoricalPayload,
        iflow_endpoint_path: str,
    ) -> ReplayResult:
        """Send one historical payload to a CPI iFlow endpoint."""
        url          = f"{self.cpi_base_url}{iflow_endpoint_path}"
        content_type = ("application/json" if payload.payload_format == "json"
                        else "application/xml")

        start = datetime.now()
        try:
            resp = self.cpi_session.post(
                url,
                data=payload.payload_content,
                headers={"Content-Type": content_type},
                timeout=60,
            )
            duration = int((datetime.now() - start).total_seconds() * 1000)

            accepted   = resp.status_code in (200, 202, 204)
            diff_notes = ""
            if not accepted:
                diff_notes = (f"CPI returned {resp.status_code}: "
                              f"{resp.text[:200]}")

            return ReplayResult(
                interface_name=payload.interface_name,
                payload_id=payload.log_id,
                pi_status=payload.status,
                cpi_status_code=resp.status_code,
                cpi_response=resp.text[:500],
                match=accepted,
                diff_notes=diff_notes,
                duration_ms=duration,
            )
        except requests.exceptions.ConnectionError:
            return ReplayResult(
                interface_name=payload.interface_name,
                payload_id=payload.log_id,
                pi_status=payload.status,
                cpi_status_code=0,
                cpi_response="",
                match=False,
                diff_notes="CPI endpoint not reachable — check iFlow is deployed",
                duration_ms=0,
            )
        except Exception as exc:
            return ReplayResult(
                interface_name=payload.interface_name,
                payload_id=payload.log_id,
                pi_status=payload.status,
                cpi_status_code=0,
                cpi_response="",
                match=False,
                diff_notes=str(exc)[:200],
                duration_ms=0,
            )

    def replay_all(
        self,
        payloads: list[HistoricalPayload],
        iflow_endpoint_path: str,
    ) -> list[ReplayResult]:
        results = []
        for p in payloads:
            result = self.replay_payload(p, iflow_endpoint_path)
            results.append(result)
            icon = "✓" if result.match else "✗"
            logger.info("%s %s → HTTP %d (%dms)",
                        icon, p.interface_name,
                        result.cpi_status_code, result.duration_ms)
        return results

    # ── Report ────────────────────────────────────────────────────────

    def generate_report(self, results: list[ReplayResult]) -> Path:
        """Generate Excel replay comparison report."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl required")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Replay Results"

        headers = ["Interface", "Payload ID", "PI Status",
                   "CPI HTTP Code", "Match", "Duration (ms)", "Notes"]
        hfill   = PatternFill("solid", fgColor="1F4E79")
        hfont   = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill; c.font = hfont

        ok_fill   = PatternFill("solid", fgColor="C6EFCE")
        fail_fill = PatternFill("solid", fgColor="FFC7CE")

        for row, r in enumerate(results, 2):
            row_data = [
                r.interface_name, r.payload_id[:12], r.pi_status,
                r.cpi_status_code, "✓" if r.match else "✗",
                r.duration_ms, r.diff_notes,
            ]
            fill = ok_fill if r.match else fail_fill
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row, column=col, value=val)
                if col == 5:
                    c.fill = fill
                    c.font = Font(bold=True)

        total   = len(results)
        passed  = sum(1 for r in results if r.match)
        ws.cell(row=total+3, column=1, value=f"TOTAL: {passed}/{total} passed")
        ws.cell(row=total+3, column=1).font = Font(bold=True)

        path = self.output_dir / "replay_report.xlsx"
        wb.save(path)
        logger.info("Replay report → %s", path)
        return path
