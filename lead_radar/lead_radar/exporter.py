"""Export results to a four-sheet xlsx pipeline tracker.

Sheets: Companies, Postings, People, SearchLinks. Data only (no formulas),
Arial, bold frozen header rows, status columns for manual triage.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .log import feature

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")
CHANNEL_FILL = PatternFill("solid", start_color="FFF2CC")


def _sheet(wb, name, headers, widths):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    return ws


def _style_row(ws, row_idx, n_cols, fill=None):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = BODY_FONT
        if fill:
            cell.fill = fill


def export_xlsx(path: str, leads: list, people: list, links: list):
    with feature("export_xlsx", path=path, companies=len(leads),
                 people=len(people), links=len(links)):
        wb = Workbook()
        wb.remove(wb.active)

        ws = _sheet(wb, "Companies",
                    ["Score", "Company", "Type", "Postings", "Best posting",
                     "URL", "Signals", "Tier", "Status", "Next action", "Notes"],
                    [8, 30, 12, 9, 45, 45, 60, 8, 14, 24, 30])
        for l in leads:
            ws.append([l.score, l.company, l.lead_type, l.n_postings,
                       l.best_posting_title, l.best_posting_url,
                       "; ".join(l.signals)[:500], "", "NEW", "", ""])
            _style_row(ws, ws.max_row, 11,
                       CHANNEL_FILL if l.lead_type == "channel" else None)

        ws = _sheet(wb, "Postings",
                    ["Score", "Company", "Title", "Location", "Posted", "Source",
                     "Query", "URL", "Score reasons"],
                    [8, 28, 45, 24, 12, 10, 32, 45, 70])
        for l in leads:
            for p in l.postings:
                ws.append([p.score, p.company, p.title, p.location, p.posted_date,
                           p.source, p.query, p.url, "; ".join(p.score_reasons)])
                _style_row(ws, ws.max_row, 9)

        ws = _sheet(wb, "People",
                    ["Role rank", "Role matched", "Company", "Name", "Title",
                     "Profile URL", "Snippet", "Status", "Connected on", "Notes"],
                    [10, 28, 26, 24, 40, 50, 60, 14, 13, 30])
        for pc in people:
            ws.append([pc.role_rank or "", pc.role_matched, pc.company, pc.name,
                       pc.title, pc.profile_url, pc.snippet, pc.status, "", ""])
            _style_row(ws, ws.max_row, 10)

        ws = _sheet(wb, "SearchLinks",
                    ["Company", "Role rank", "Role", "Google X-ray URL"],
                    [30, 10, 32, 110])
        for company, rank, label, url in links:
            ws.append([company, rank, label, url])
            _style_row(ws, ws.max_row, 4)

        wb.save(path)
    return path
