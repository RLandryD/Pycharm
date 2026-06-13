"""Reverse Migration Assessment export — real CPI/PI packages IN, an
SAP-Migration-Assessment-format xlsx OUT.

Why: the forward path (parse SAP's MA xlsx → assessments) covers clients who
ran PIMAS. This module covers the opposite, common case: we HAVE the package
bytes (tenant pull or PI export) and want to hand the client an assessment in
the exact format their architects already know — Executive Summary dashboard,
Scenario Evaluation grid, Rules Log — without anyone running SAP's tool.

Layout is replicated cell-for-cell from a real SAP export (verified against
a 169-interface reference file):
  · Executive Summary — title B2, subtitle B3, 'Sizing Dashboard Summary' B5,
    header row 6 (Metric / KPI Value / Percentage / Description), metric rows
    7-11 (Total ICOs, Ready/Low, Adjustment/Medium, Evaluation/High, Total
    Effort 'X.X Hrs').
  · Scenario Evaluation — header row 4, cols B-J: ICO Technical ID, Sender
    System, Receiver System, Sender Adapter, Receiver Adapter, Complexity
    Group, Migration Status, Rule Weight, Est. Effort ('X.X Hrs' string).
  · Rules Log — header row 4, cols B-F: Triggered Rule ID, Target ICO
    Scenario Affected, Complexity, Identified Asset String / Context,
    Assessment Technical Note & Remediation Strategy.

Sizes map to the export vocabulary: S→Low/Ready to Migrate,
M→Medium/Adjustment Required, L/XL→High/Evaluation Required. Effort cells
take the midpoint of the engine's per-flow effort band.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field

logger = logging.getLogger("reporter.ma_export")

_GROUP = {"S": "Low", "M": "Medium", "L": "High", "XL": "High"}
_STATUS = {"S": "Ready to Migrate", "M": "Adjustment Required",
           "L": "Evaluation Required", "XL": "Evaluation Required"}

_HDR_FILL = "FF1F4E79"      # dark blue header band (mirrors the SAP export)
_HDR_FONT_COLOR = "FFFFFFFF"


@dataclass
class MAExportRow:
    """One assessed flow, ready for the Scenario Evaluation grid."""
    ico_id: str
    sender_system: str = "N/A"
    receiver_system: str = "N/A"
    sender_adapter: str = ""
    receiver_adapter: str = ""
    size: str = "M"
    weight: int = 0
    effort_hours: float = 0.0
    fired_rules: list = field(default_factory=list)   # FiredRule objects


def result_to_row(result, ico_id: str = "",
                  sender_system: str = "N/A",
                  receiver_system: str = "N/A") -> MAExportRow:
    """Adapt a SAPComplexityEngine ComplexityResult to an export row."""
    sig = getattr(result, "signals", {}) or {}
    lo = float(getattr(result, "effort_hours_low", 0) or 0)
    hi = float(getattr(result, "effort_hours_high", 0) or 0)
    return MAExportRow(
        ico_id=ico_id or getattr(result, "name", "") or "ICO",
        sender_system=sender_system,
        receiver_system=receiver_system,
        sender_adapter=str(sig.get("sender_adapter", "") or ""),
        receiver_adapter=str(sig.get("receiver_adapter", "") or ""),
        size=getattr(result, "size", "M") or "M",
        weight=int(getattr(result, "total_weight", 0) or 0),
        effort_hours=round((lo + hi) / 2.0, 1) if (lo or hi) else 0.0,
        fired_rules=list(getattr(result, "fired_rules", []) or []),
    )


def build_ma_xlsx(rows: list, out_path: str,
                  title: str = "SAP Integration Suite Migration Assessment",
                  subtitle: str = "") -> str:
    """Write the 3-sheet MA-format workbook. Returns out_path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor=_HDR_FILL)
    hdr_font = Font(bold=True, color=_HDR_FONT_COLOR)
    title_font = Font(bold=True, size=14)
    sub_font = Font(italic=True, size=10)

    n = len(rows)
    low = sum(1 for r in rows if r.size == "S")
    med = sum(1 for r in rows if r.size == "M")
    high = n - low - med
    total_hrs = round(sum(r.effort_hours for r in rows), 1)

    def pct(x):
        return round(x / n, 3) if n else 0

    wb = Workbook()

    # ── Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"
    ws["B2"] = title
    ws["B2"].font = title_font
    ws["B3"] = subtitle or (f"Package Scan Analysis ({n} Interface"
                            f"{'s' if n != 1 else ''})")
    ws["B3"].font = sub_font
    ws["B5"] = "Sizing Dashboard Summary"
    ws["B5"].font = Font(bold=True, size=12)
    for col, txt in zip("BCDE", ("Metric", "KPI Value", "Percentage",
                                 "Description / Primary Action Items")):
        c = ws[f"{col}6"]
        c.value, c.font, c.fill = txt, hdr_font, hdr_fill
    metrics = [
        ("Total Extracted ICOs", n, "100%",
         "Total integration flows assessed from real package bytes "
         "(adapter, mapping, script, and structure signals)."),
        ("Ready to Migrate (Pure Low)", low, pct(low),
         "Standard interface pipelines eligible for direct automated "
         "conversion templates."),
        ("Adjustment Required (Medium)", med, pct(med),
         "Standard paths requiring explicit manual parameterization or "
         "endpoint reconfiguration."),
        ("Evaluation Required (High & Trapped Low)", high, pct(high),
         "CRITICAL BLOCKERS: legacy custom code, heavy transformation "
         "chains, or architectural dependencies."),
        ("Total Estimated Effort Portfolio", f"{total_hrs} Hrs", "-",
         "Sum total calculated engineering resource hours for migration, "
         "adjustment, and refactoring testing phases."),
    ]
    for i, (m, v, p, d) in enumerate(metrics, start=7):
        ws[f"B{i}"], ws[f"C{i}"], ws[f"D{i}"], ws[f"E{i}"] = m, v, p, d
    for col, w in (("B", 38), ("C", 12), ("D", 12), ("E", 80)):
        ws.column_dimensions[col].width = w

    # ── Scenario Evaluation ──
    ws2 = wb.create_sheet("Scenario Evaluation")
    headers = ["ICO Technical ID", "Sender System", "Receiver System",
               "Sender Adapter", "Receiver Adapter", "Complexity Group",
               "Migration Status", "Rule Weight", "Est. Effort"]
    for j, h in enumerate(headers):
        c = ws2.cell(row=4, column=2 + j, value=h)
        c.font, c.fill = hdr_font, hdr_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, r in enumerate(rows, start=5):
        vals = [r.ico_id, r.sender_system, r.receiver_system,
                r.sender_adapter, r.receiver_adapter,
                _GROUP.get(r.size, "Medium"), _STATUS.get(r.size, ""),
                r.weight, f"{r.effort_hours} Hrs"]
        for j, v in enumerate(vals):
            ws2.cell(row=i, column=2 + j, value=v)
    widths = (46, 18, 18, 16, 16, 16, 20, 12, 12)
    for j, w in enumerate(widths):
        ws2.column_dimensions[get_column_letter(2 + j)].width = w

    # ── Rules Log ──
    ws3 = wb.create_sheet("Rules Log")
    headers3 = ["Triggered Rule ID", "Target ICO Scenario Affected",
                "Complexity", "Identified Asset String / Context",
                "Assessment Technical Note & Remediation Strategy"]
    for j, h in enumerate(headers3):
        c = ws3.cell(row=4, column=2 + j, value=h)
        c.font, c.fill = hdr_font, hdr_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    rr = 5
    for r in rows:
        for f in r.fired_rules:
            asset = (getattr(f, "signal_note", "") or
                     getattr(f, "matched_value", "") or "")
            note = (f"Rule band '{getattr(f, 'matched_value', '')}' matched "
                    f"(weight +{getattr(f, 'weight', 0)}). Review the "
                    f"corresponding CPI artifacts during migration design.")
            vals = [getattr(f, "rule", ""), r.ico_id,
                    _GROUP.get(r.size, "Medium"), asset, note]
            for j, v in enumerate(vals):
                ws3.cell(row=rr, column=2 + j, value=v)
            rr += 1
    widths3 = (34, 46, 12, 34, 80)
    for j, w in enumerate(widths3):
        ws3.column_dimensions[get_column_letter(2 + j)].width = w

    wb.save(out_path)
    logger.info("MA-format export written → %s (%d ICOs, %d rule rows, "
                "%.1f hrs)", out_path, n, rr - 5, total_hrs)
    return out_path
