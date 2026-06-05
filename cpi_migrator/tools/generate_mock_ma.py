"""Generate a faithful MOCK SAP Migration Assessment export (.xlsx).

This is a TEST FIXTURE, not real SAP output. It matches the exact 3-sheet
structure the parser (intake/sap_ma_parser.py) reads from a real MA export:

    Executive Summary   — Metric / KPI Value
    Rules Log           — Triggered Rule / Target ICO / Asset String / Technical Note
    Scenario Evaluation — ICO Technical ID / Sender System / Receiver System /
                          Sender Adapter / Receiver Adapter / Mapping Types Found /
                          Migration Status / Rule Weight / Estimated Effort

All weights, sizes, categories and effort hours are derived from SAP's REAL
extracted model (analyzer/data/*.json), so the numbers are internally
consistent with the engine. The PER-INTERFACE problem mix is a realistic
*construction* (we have no real evaluation to copy), seeded for variety:
trivial pass-throughs, graphical-mapping mediums, and genuine Evaluate
monsters (ABAP mapping, ccBPM, custom adapter modules, Java mapping).

Honest scope: this validates the Mode-1 consumption path + parser + the whole
downstream consumer chain (proposals/estimates/reports) end-to-end. It does
NOT validate "which rule should fire on a real interface" — only real PO/MA
data can do that.
"""
from __future__ import annotations

import json
import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

_DATA = Path(__file__).parent.parent / "analyzer" / "data"
_CATALOG = json.load(open(_DATA / "sap_ma_rule_catalog.json"))
_W2S = json.load(open(_DATA / "sap_ma_weight_to_size.json"))
_EFFORT = json.load(open(_DATA / "sap_ma_effort_table.json"))["table"]

_CAT_LONG = {"Migrate": "Ready to migrate", "Adapt": "Adjustments required",
             "Evaluate": "Evaluation required"}
_CAT_SHORT = {"Ready to migrate": "Migrate", "Adjustments required": "Adapt",
              "Adjustment required": "Adapt",  # tolerate singular variant
              "Evaluation required": "Evaluate"}


def _weight_to_size(w: int) -> str:
    for row in _W2S:
        if row["FromWeight"] <= w <= row["ToWeight"]:
            return row["Size"]
    return "XL"


def _effort_mid(size: str, cat_short: str) -> float:
    cell = _EFFORT.get(size, {}).get(cat_short, [0, 0])
    return round((cell[0] + cell[1]) / 2, 1)


# Real systems / adapters for realistic metadata.
_SYSTEMS = ["S4HANA_PRD", "ECC_PRD", "SFSF_PRD", "ARIBA_PRD", "SUCCESSFACTORS",
            "MDG_PRD", "CPI_TENANT", "OPENTEXT", "SALESFORCE", "WORKDAY",
            "THIRD_PARTY_SFTP", "BANK_HOST", "LEGACY_AS400", "CONCUR"]
_ADAPTERS = ["IDoc", "SOAP", "HTTP", "SFTP", "File", "JDBC", "RFC", "REST",
             "OData", "Mail", "JMS", "AS2", "SFSF"]

# Interface "archetypes" — each is a realistic problem profile that fires a
# specific set of REAL rules. (rule_name, characteristic-or-'', count_or_value)
_ARCHETYPES = {
    "trivial_passthrough": [  # IDoc->File, no mapping  -> S / Migrate
        ("ICOReceivers", "", "2"),
        ("ICOOperationCount", "RECEIVER_IF", "2"),
    ],
    "simple_graphical": [  # one graphical mapping, a couple UDFs -> S/M Migrate
        ("MappingType", "RECEIVER_IF", "GMM"),
        ("GMMCustomUDFUsageCount", "RECEIVER_IF", "5"),
        ("OMStepCount", "RECEIVER_IF", "3"),
        ("ICOReceivers", "", "2"),
    ],
    "xslt_heavy": [  # multiple XSLT deps -> M, Migrate
        ("MappingType", "RECEIVER_IF", "XSL"),
        ("XSLTDependenciesCount", "RECEIVER_IF", "5"),
        ("XSLTDependenciesCount", "EXT_RCV_DET", "5"),
        ("OMStepCount", "RECEIVER_IF", "4"),
    ],
    "udf_lookup_heavy": [  # many UDFs + RFC/JDBC lookups -> M, Migrate/Adapt
        ("MappingType", "RECEIVER_IF", "GMM"),
        ("GMMCustomUDFUsageCount", "RECEIVER_IF", "150"),
        ("GMMRFCLookupCount", "RECEIVER_IF", "3"),
        ("GMMJDBCLookUpCount", "RECEIVER_IF", "2"),
        ("OMStepCount", "RECEIVER_IF", "6"),
    ],
    "jms_qos": [  # JMS adapter -> Adapt
        ("ReceiverAdapterType", "", "JMS"),
        ("SenderAdapterType", "", "File"),  # File sender is Adapt-category
        ("MappingType", "RECEIVER_IF", "GMM"),
        ("MappingType", "EXT_RCV_DET", "GMM"),
    ],
    "file_cc_adapt": [  # File content conversion + File adapter -> Adapt
        ("SenderAdapterType", "", "File"),
        ("MappingType", "RECEIVER_IF", "GMM"),
        ("GMMCustomUDFUsageCount", "RECEIVER_IF", "80"),
        ("GMMCustomUDFUsageCount", "EXT_RCV_DET", "80"),
        ("OMStepCount", "RECEIVER_IF", "4"),
    ],
    "abap_mapping": [  # ABAP mapping -> EVALUATE (both chars -> bigger)
        ("MappingType", "RECEIVER_IF", "ABAP"),
        ("MappingType", "EXT_RCV_DET", "ABAP"),
        ("GMMCustomUDFUsageCount", "RECEIVER_IF", "120"),
        ("GMMCustomUDFUsageCount", "EXT_RCV_DET", "120"),
        ("OMStepCount", "RECEIVER_IF", "5"),
        ("OMStepCount", "EXT_RCV_DET", "5"),
    ],
    "ccbpm": [  # classical ccBPM -> EVALUATE
        ("ccBPM", "", "true"),
        ("MappingType", "RECEIVER_IF", "GMM"),
        ("OMStepCount", "RECEIVER_IF", "5"),
        ("OMStepCount", "EXT_RCV_DET", "5"),
        ("GMMCustomUDFUsageCount", "RECEIVER_IF", "150"),
        ("GMMCustomUDFUsageCount", "EXT_RCV_DET", "150"),
        ("XSLTDependenciesCount", "RECEIVER_IF", "5"),
    ],
    "java_mapping": [  # Java mapping + Java BPM -> EVALUATE
        ("MappingType", "RECEIVER_IF", "Java"),
        ("MappingType", "EXT_RCV_DET", "Java"),
        ("JavaBPM", "", "true"),
        ("OMStepCount", "RECEIVER_IF", "4"),
        ("XSLTDependenciesCount", "RECEIVER_IF", "5"),
        ("XSLTDependenciesCount", "EXT_RCV_DET", "5"),
    ],
    "custom_adapter_module": [  # custom adapter module -> EVALUATE (XL monster)
        ("SenderAdapterType", "", "RFC"),
        ("MappingType", "RECEIVER_IF", "ABAP"),
        ("MappingType", "EXT_RCV_DET", "ABAP"),
        ("GMMCustomUDFUsageCount", "RECEIVER_IF", "250"),
        ("GMMCustomUDFUsageCount", "EXT_RCV_DET", "250"),
        ("GMMRFCLookupCount", "RECEIVER_IF", "5"),
        ("GMMRFCLookupCount", "EXT_RCV_DET", "5"),
        ("GMMJDBCLookUpCount", "RECEIVER_IF", "3"),
        ("XSLTDependenciesCount", "RECEIVER_IF", "5"),
        ("XSLTDependenciesCount", "EXT_RCV_DET", "5"),
        ("OMStepCount", "RECEIVER_IF", "8"),
        ("OMStepCount", "EXT_RCV_DET", "8"),
        ("ccBPM", "", "true"),
    ],
}

# ---------------------------------------------------------------------------
# Coverage-matrix construction
# ---------------------------------------------------------------------------
# Reusable REAL bands as stackable "weight blocks". Each is (rule, char, token).
# Migrate fillers (stay Ready to migrate, push weight up):
_MIG_70 = ("GMMCustomUDFUsageCount", "EXT_RCV_DET", "250")   # 70, Migrate
_MIG_70b = ("GMMRFCLookupCount", "RECEIVER_IF", "250")       # 70, Migrate
_MIG_70c = ("GMMJDBCLookUpCount", "EXT_RCV_DET", "250")      # 70, Migrate
_MIG_50 = ("GMMCustomUDFUsageCount", "RECEIVER_IF", "250")   # 50, Migrate
_MIG_50b = ("XSLTDependenciesCount", "RECEIVER_IF", "5")     # 50, Migrate
_MIG_50c = ("XSLTDependenciesCount", "EXT_RCV_DET", "5")     # 50, Migrate
_MIG_15 = ("OMStepCount", "EXT_RCV_DET", "5")                # 50? -> check; use small
_MIG_5 = ("MappingType", "RECEIVER_IF", "GMM")              # 5, Migrate
_MIG_2 = ("ICOReceivers", "", "2")                          # 1, Migrate
# Adapt driver (worst-case Adapt; weight 50):
_ADAPT_50 = ("SenderJMSTransportProtocol", "ADAPTER", "SonicMQ")  # 50, Adapt
_ADAPT_15 = ("SenderAdapterType", "", "File")                    # 15, Adapt
# Evaluate driver (worst-case Evaluate; weight 50):
_EVAL_50 = ("ccBPM", "", "true")                            # 50, Evaluate
_EVAL_50b = ("MappingType", "RECEIVER_IF", "ABAP")          # 50, Evaluate

# Size band centres we aim for: S~80, M~250, L~430, XL~560/max.
# Each matrix entry: (cell_label, size, category, [blocks], note)
def _matrix_specs():
    specs = []
    # ---- S row (weight 1-150) ----
    specs.append(("S_Migrate", [_MIG_5, _MIG_2, ("OMStepCount", "RECEIVER_IF", "3")]))
    specs.append(("S_Adapt",   [_ADAPT_15, _MIG_5, _MIG_2]))             # File adapter -> Adapt, small
    specs.append(("S_Evaluate",[_EVAL_50, _MIG_5, _MIG_2]))             # ccBPM but low weight -> S/Evaluate
    # ---- M row (151-350) ----
    specs.append(("M_Migrate", [_MIG_70, _MIG_70b, _MIG_50, _MIG_50b]))  # ~240, Migrate
    specs.append(("M_Adapt",   [_ADAPT_50, _MIG_70, _MIG_70b, _MIG_50])) # ~240, Adapt worst-case
    specs.append(("M_Evaluate",[_EVAL_50, _MIG_70, _MIG_70b, _MIG_50]))  # ~240, Evaluate worst-case
    # ---- L row (351-500) ----
    Lfill = [_MIG_70, _MIG_70b, _MIG_70c, _MIG_50, _MIG_50b, _MIG_50c]   # 360 Migrate
    specs.append(("L_Migrate", Lfill))
    specs.append(("L_Adapt",   [_ADAPT_50] + Lfill[:-1]))               # 50 + 310 = 360, Adapt
    specs.append(("L_Evaluate",[_EVAL_50] + Lfill[:-1]))                # 360, Evaluate
    # ---- XL row (501+) — ONE max-weight monster (Evaluate) ----
    XLfill = [_MIG_70, _MIG_70b, _MIG_70c, _MIG_50, _MIG_50b, _MIG_50c,
              ("GMMCustomFuncLibUsageCount", "RECEIVER_IF", "50"),  # +20
              ("GMMValueMapping", "", "5"),                         # + small
              ("OMStepCount", "RECEIVER_IF", "5"),
              ("OMStepCount", "EXT_RCV_DET", "5"),
              ("ICOOperationCount", "RECEIVER_IF", "5"),
              ("ICOOperationCount", "EXT_RCV_DET", "5")]
    specs.append(("XL_Migrate", XLfill + [_MIG_50, _MIG_5, _MIG_2,
                  ("GMMJDBCLookUpCount", "RECEIVER_IF", "250")]))
    specs.append(("XL_Adapt",   [_ADAPT_50, _ADAPT_15] + XLfill))
    # The single max-weight interface: everything heavy + Evaluate.
    specs.append(("XL_Evaluate_MAX", [_EVAL_50, _EVAL_50b, _ADAPT_50,
                  _MIG_70, _MIG_70b, _MIG_70c, _MIG_50, _MIG_50b, _MIG_50c,
                  ("GMMCustomFuncLibUsageCount", "RECEIVER_IF", "50"),
                  ("OMStepCount", "RECEIVER_IF", "5"),
                  ("OMStepCount", "EXT_RCV_DET", "5"),
                  ("ICOOperationCount", "RECEIVER_IF", "5")]))
    return specs


def _band_weight(rule, characteristic, token):
    """Resolve a real catalog band weight+category for an archetype entry."""
    full = f"MAIN_{rule}"
    bands = _CATALOG.get(full, [])
    # numeric token -> range band; else value/list membership
    try:
        n = int(token)
        is_num = True
    except ValueError:
        is_num = False
    for b in bands:
        if characteristic and b.get("characteristic") and b["characteristic"] != characteristic:
            continue
        mv = b.get("match_value", "")
        if is_num and "," in mv:
            lo, hi = mv.split(",")[0], mv.split(",")[-1]
            try:
                if int(lo) <= n <= int(hi):
                    return int(b["weight"]), _CAT_SHORT.get(b["category"], "Migrate")
            except ValueError:
                continue
        elif not is_num:
            toks = [t.strip().lower() for t in mv.split(",")]
            if token.lower() == mv.lower() or token.lower() in toks:
                return int(b["weight"]), _CAT_SHORT.get(b["category"], "Migrate")
    return 0, "Migrate"


def generate(path: str, n: int = 45, seed: int = 42) -> dict:
    random.seed(seed)
    cat_rank = {"Migrate": 0, "Adapt": 1, "Evaluate": 2}
    rank_cat = {0: "Migrate", 1: "Adapt", 2: "Evaluate"}

    interfaces = []
    rules_log = []

    def build_one(ico_id, blocks):
        total_w = 0
        worst = 0
        mapping_types = set()
        for (rule, ch, token) in blocks:
            w, cat = _band_weight(rule, ch, token)
            total_w += w
            worst = max(worst, cat_rank[cat])
            if rule == "MappingType":
                mapping_types.add(token)
            rules_log.append({
                "rule": rule, "ico": ico_id,
                "asset": token if token != "true" else "present",
                "note": _NOTE.get(rule, "Review required during migration."),
            })
        category = rank_cat[worst]
        size = _weight_to_size(total_w)
        effort = _effort_mid(size, category)
        return {
            "id": ico_id,
            "sender_sys": random.choice(_SYSTEMS),
            "receiver_sys": random.choice(_SYSTEMS),
            "sender_adapter": random.choice(_ADAPTERS),
            "receiver_adapter": random.choice(_ADAPTERS),
            "mapping_types": " + ".join(sorted(mapping_types)) if mapping_types else "None",
            "status": _CAT_LONG[category],
            "weight": total_w, "size": size, "effort": effort,
            "_target": None,
        }

    # 1) One interface per matrix cell (the coverage guarantee).
    for i, (label, blocks) in enumerate(_matrix_specs(), 1):
        rec = build_one(f"ICO_{i:03d}_{label}", blocks)
        rec["_target"] = label
        interfaces.append(rec)

    # 2) Pad with realistic everyday interfaces to reach n (variety, not edges).
    pad_blocks = [
        [_MIG_5, _MIG_2, ("OMStepCount", "RECEIVER_IF", "3")],          # tiny
        [_MIG_50, _MIG_5, ("GMMCustomUDFUsageCount", "RECEIVER_IF", "5")],
        [_MIG_50b, _MIG_5, _MIG_2],
        [("MappingType", "RECEIVER_IF", "GMM"), ("GMMCustomUDFUsageCount", "RECEIVER_IF", "50")],
        [_ADAPT_15, _MIG_5, _MIG_2],
    ]
    idx = len(interfaces)
    while len(interfaces) < n:
        idx += 1
        blocks = random.choice(pad_blocks)
        interfaces.append(build_one(f"ICO_{idx:03d}_STD", blocks))

    _write_xlsx(path, interfaces, rules_log)

    from collections import Counter
    cells = {}
    for x in interfaces:
        if x["_target"]:
            cells[x["_target"]] = (x["size"], _CAT_SHORT[x["status"]], x["weight"])
    return {
        "count": len(interfaces),
        "sizes": dict(Counter(x["size"] for x in interfaces)),
        "categories": dict(Counter(_CAT_SHORT[x["status"]] for x in interfaces)),
        "total_effort": round(sum(x["effort"] for x in interfaces), 1),
        "weight_range": (min(x["weight"] for x in interfaces),
                         max(x["weight"] for x in interfaces)),
        "matrix_cells": cells,
    }


_NOTE = {
    "ccBPM": "Classical ccBPM detected — no direct CPI equivalent; redesign as iFlow orchestration.",
    "JavaBPM": "Java BPM in use — requires redesign in Cloud Integration.",
    "MappingType": "Mapping must be re-implemented in CPI.",
    "GMMCustomUDFUsageCount": "Custom UDFs must be ported to Groovy/CPI functions.",
    "GMMRFCLookupCount": "RFC lookups require a corresponding CPI receiver/Request-Reply.",
    "GMMJDBCLookUpCount": "JDBC lookups require a CPI JDBC receiver.",
    "XSLTDependenciesCount": "XSLT dependencies must be bundled into the CPI artifact.",
    "ReceiverAdapterType": "Receiver adapter type requires CPI equivalent configuration.",
    "SenderAdapterType": "Sender adapter type requires CPI equivalent configuration.",
    "OMStepCount": "Operation mapping steps to be reproduced in CPI.",
    "ICOReceivers": "Multiple receivers — consider multicast/routing in CPI.",
    "ICOOperationCount": "Service interface operations to be migrated.",
}


def _write_xlsx(path, interfaces, rules_log):
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)

    def style_header(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row, c)
            cell.fill = hdr_fill
            cell.font = hdr_font

    # --- Sheet 1: Executive Summary -------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.cell(1, 1, "SAP Integration Suite — Migration Assessment (MOCK FIXTURE)")
    ws1.cell(1, 1).font = Font(bold=True, size=14)
    ws1.cell(2, 1, "Generated test data — not a real SAP evaluation")
    ws1.cell(2, 1).font = Font(italic=True, color="888888")
    from collections import Counter
    cats = Counter(_CAT_SHORT[x["status"]] for x in interfaces)
    total_effort = round(sum(x["effort"] for x in interfaces), 1)
    ws1.cell(4, 1, "Metric"); ws1.cell(4, 2, "KPI Value")
    style_header(ws1, 4, 2)
    rows = [
        ("Total Extracted ICOs", len(interfaces)),
        ("Ready to migrate", cats.get("Migrate", 0)),
        ("Adjustment required", cats.get("Adapt", 0)),
        ("Evaluation required", cats.get("Evaluate", 0)),
        ("Total Estimated Effort", f"{total_effort} Hrs"),
    ]
    for i, (m, v) in enumerate(rows, start=5):
        ws1.cell(i, 1, m); ws1.cell(i, 2, v)
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 18

    # --- Sheet 2: Rules Log ---------------------------------------------
    ws2 = wb.create_sheet("Rules Log")
    headers2 = ["Triggered Rule", "Target ICO", "Asset String", "Technical Note"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(1, c, h)
    style_header(ws2, 1, len(headers2))
    for r, row in enumerate(rules_log, start=2):
        ws2.cell(r, 1, row["rule"]); ws2.cell(r, 2, row["ico"])
        ws2.cell(r, 3, row["asset"]); ws2.cell(r, 4, row["note"])
    for col, w in zip("ABCD", (32, 28, 22, 60)):
        ws2.column_dimensions[col].width = w

    # --- Sheet 3: Scenario Evaluation -----------------------------------
    ws3 = wb.create_sheet("Scenario Evaluation")
    headers3 = ["ICO Technical ID", "Sender System", "Receiver System",
                "Sender Adapter", "Receiver Adapter", "Mapping Types Found",
                "Migration Status", "Rule Weight", "Estimated Effort"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(1, c, h)
    style_header(ws3, 1, len(headers3))
    for r, x in enumerate(interfaces, start=2):
        ws3.cell(r, 1, x["id"]); ws3.cell(r, 2, x["sender_sys"])
        ws3.cell(r, 3, x["receiver_sys"]); ws3.cell(r, 4, x["sender_adapter"])
        ws3.cell(r, 5, x["receiver_adapter"]); ws3.cell(r, 6, x["mapping_types"])
        ws3.cell(r, 7, x["status"]); ws3.cell(r, 8, x["weight"])
        ws3.cell(r, 9, f"{x['effort']} Hrs")
    for col, w in zip("ABCDEFGHI", (26, 16, 16, 14, 16, 20, 20, 12, 16)):
        ws3.column_dimensions[col].width = w

    wb.save(path)


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "mock_sap_ma_export.xlsx"
    n = int(sys.argv[2]) if len(sys.argv) > 2 else 45
    summary = generate(out, n=n)
    print(f"Wrote {out}")
    print(json.dumps(summary, indent=2))
