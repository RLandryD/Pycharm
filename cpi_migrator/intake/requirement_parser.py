"""
intake/requirement_parser.py

Parses a technical requirement document (Excel, Word, or plain text)
and pre-populates InterfaceConfig + InterfaceRecord fields.

Two modes:
  1. Structured Excel template  — deterministic, parses named columns
  2. Free-text / Word document  — uses Claude API to extract fields as JSON

Output: list[RequirementResult] each containing a pre-filled
        InterfaceRecord + InterfaceConfig ready for the workbench.
"""
from __future__ import annotations

import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Output model
# ---------------------------------------------------------------------------

@dataclass
class RequirementResult:
    """One integration requirement parsed from the input document."""
    # Pre-filled record (for analyzer + scaffolder)
    name: str
    sender_system: str        = ""
    sender_adapter: str       = "HTTPS"
    receiver_system: str      = ""
    receiver_adapter: str     = "HTTPS"
    namespace: str            = ""
    description: str          = ""
    message_interface: str    = ""
    mapping_program: str      = ""

    # Config fields
    target_id: str            = "s4hana_cloud"
    sender_address: str       = ""
    receiver_address: str     = ""
    sender_auth_method: str   = "Basic"
    receiver_auth_method: str = "OAuth2 Client Credentials"
    is_async: bool            = False
    message_format: str       = "XML"
    scheduler_cron: str       = ""
    business_process: str     = ""

    # Confidence: how much was auto-filled vs. needs review
    confidence: float         = 0.0    # 0.0–1.0
    needs_review: list[str]   = field(default_factory=list)
    raw_text: str             = ""

    def to_interface_record(self):
        from extractor.pi_extractor import InterfaceRecord
        return InterfaceRecord(
            id=re.sub(r"[^\w]", "_", self.name),
            name=self.name,
            namespace=self.namespace,
            software_component="",
            sender_system=self.sender_system,
            receiver_system=self.receiver_system,
            sender_adapter=self.sender_adapter,
            receiver_adapter=self.receiver_adapter,
            message_interface=self.message_interface,
            mapping_program=self.mapping_program or None,
            description=self.description,
        )

    def to_interface_config(self):
        from models.interface_config import InterfaceConfig, AuthConfig, ConnectivityConfig, MessageConfig
        cfg = InterfaceConfig(
            interface_name=self.name,
            target_id=self.target_id,
            sender_adapter=self.sender_adapter,
            receiver_adapter=self.receiver_adapter,
        )
        cfg.sender_connectivity.address  = self.sender_address
        cfg.receiver_connectivity.address = self.receiver_address
        cfg.sender_auth.method           = self.sender_auth_method
        cfg.receiver_auth.method         = self.receiver_auth_method
        cfg.message.is_async             = self.is_async
        cfg.message.format               = self.message_format
        cfg.message.mapping_program      = self.mapping_program
        cfg.message.namespace            = self.namespace
        cfg.runtime.scheduler_cron       = self.scheduler_cron
        cfg.manual_steps                 = [f"Review: {f}" for f in self.needs_review]
        return cfg


# ---------------------------------------------------------------------------
# Structured Excel parser
# ---------------------------------------------------------------------------

# Column name aliases (case-insensitive, spaces/underscores stripped)
EXCEL_COLUMN_MAP = {
    "integrationname":     "name",
    "interfacename":       "name",
    "name":                "name",
    "sendersystem":        "sender_system",
    "source":              "sender_system",
    "sourcesystem":        "sender_system",
    "senderadapter":       "sender_adapter",
    "sourceadapter":       "sender_adapter",
    "receiversystem":      "receiver_system",
    "target":              "receiver_system",
    "targetsystem":        "receiver_system",
    "destination":         "receiver_system",
    "receiveradapter":     "receiver_adapter",
    "targetadapter":       "receiver_adapter",
    "namespace":           "namespace",
    "description":         "description",
    "businessprocess":     "business_process",
    "process":             "business_process",
    "messageformat":       "message_format",
    "format":              "message_format",
    "async":               "is_async",
    "synchronous":         "is_async",
    "scheduler":           "scheduler_cron",
    "cron":                "scheduler_cron",
    "targetid":            "target_id",
    "destinationtype":     "target_id",
    "senderaddress":       "sender_address",
    "sourceurl":           "sender_address",
    "receiveraddress":     "receiver_address",
    "targeturl":           "receiver_address",
    "mappingprogram":      "mapping_program",
    "mapping":             "mapping_program",
    "messageinterface":    "message_interface",
}

TARGET_ALIASES = {
    "s4":             "s4hana_cloud",
    "s4hana":         "s4hana_cloud",
    "s4cloud":        "s4hana_cloud",
    "s4op":           "s4hana_op",
    "s4onpremise":    "s4hana_op",
    "ariba":          "ariba",
    "successfactors": "successfactors",
    "sf":             "successfactors",
    "sfsf":           "successfactors",
    "btp":            "btp",
    "fieldglass":     "fieldglass",
    "concur":         "concur",
}


class ExcelRequirementParser:
    """Parses a structured Excel requirement template."""

    def parse(self, file_path: str) -> list[RequirementResult]:
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl required: pip install openpyxl")

        wb   = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Map header → field name
        headers = [
            EXCEL_COLUMN_MAP.get(
                str(h or "").strip().lower().replace(" ", "").replace("_", ""),
                None
            )
            for h in rows[0]
        ]

        results = []
        for row in rows[1:]:
            if not any(row):
                continue
            data = {}
            for col_idx, field_name in enumerate(headers):
                if field_name and col_idx < len(row) and row[col_idx] is not None:
                    data[field_name] = str(row[col_idx]).strip()

            name = data.get("name", "").strip()
            if not name:
                continue

            # Normalise target_id
            raw_target = data.get("target_id", "s4hana_cloud").lower().replace(" ", "")
            target_id  = TARGET_ALIASES.get(raw_target, raw_target)

            # Normalise is_async
            raw_async = data.get("is_async", "false").lower()
            is_async  = raw_async in ("true", "yes", "1", "async", "asynchronous")

            needs_review = []
            confidence   = 0.5

            result = RequirementResult(
                name=name,
                sender_system=data.get("sender_system", ""),
                sender_adapter=data.get("sender_adapter", "HTTPS"),
                receiver_system=data.get("receiver_system", ""),
                receiver_adapter=data.get("receiver_adapter", "HTTPS"),
                namespace=data.get("namespace", ""),
                description=data.get("description", ""),
                message_interface=data.get("message_interface", ""),
                mapping_program=data.get("mapping_program", ""),
                target_id=target_id,
                sender_address=data.get("sender_address", ""),
                receiver_address=data.get("receiver_address", ""),
                is_async=is_async,
                message_format=data.get("message_format", "XML"),
                scheduler_cron=data.get("scheduler_cron", ""),
                business_process=data.get("business_process", ""),
                confidence=confidence,
                needs_review=needs_review,
            )

            if not result.sender_system:
                needs_review.append("sender_system not specified")
            if not result.receiver_system:
                needs_review.append("receiver_system not specified")
            if not result.sender_address:
                needs_review.append("sender_address/URL not specified")
            if not result.receiver_address:
                needs_review.append("receiver_address/URL not specified")

            result.confidence = 1.0 - (len(needs_review) * 0.1)
            results.append(result)

        logger.info("Parsed %d requirements from Excel", len(results))
        return results


# ---------------------------------------------------------------------------
# Free-text / AI parser (uses Claude API)
# ---------------------------------------------------------------------------

class AIRequirementParser:
    """
    Parses free-text technical requirement documents using Claude API.
    Accepts plain text, copied Word content, or uploaded .txt/.docx content.
    """

    SYSTEM_PROMPT = """You are an SAP integration architect. 
Extract integration requirements from the provided document and return ONLY a JSON array.
Each element represents one integration interface with these fields:
{
  "name": "interface name or identifier",
  "sender_system": "source system name",
  "sender_adapter": "adapter type: HTTPS/SOAP/IDoc/RFC/File/SFTP/FTP/JDBC/JMS/AS2/OData/REST",
  "receiver_system": "target system name",
  "receiver_adapter": "adapter type (same options)",
  "description": "brief description of what this integration does",
  "is_async": true or false,
  "message_format": "XML/JSON/IDoc/CSV/Binary",
  "target_id": "s4hana_cloud/s4hana_op/ariba/successfactors/btp/fieldglass/concur",
  "sender_address": "URL or host if mentioned",
  "receiver_address": "URL or host if mentioned",
  "mapping_program": "mapping name if mentioned",
  "namespace": "XML namespace if mentioned",
  "scheduler_cron": "cron expression if scheduled, empty otherwise",
  "business_process": "business process name"
}
Return ONLY the JSON array. No markdown, no explanation, no code fences."""

    def parse(self, text: str) -> list[RequirementResult]:
        """Parse free text using Claude API."""
        try:
            import requests as req
            response = req.post(
                "https://api.anthropic.com/v1/messages",
                headers={"Content-Type": "application/json"},
                json={
                    "model": "claude-sonnet-4-20250514",
                    "max_tokens": 2000,
                    "system": self.SYSTEM_PROMPT,
                    "messages": [{"role": "user", "content": text}],
                },
                timeout=30,
            )
            response.raise_for_status()
            data    = response.json()
            content = next(
                (b["text"] for b in data.get("content", []) if b.get("type") == "text"), ""
            )
            # Strip any accidental markdown fences
            content = re.sub(r"```(?:json)?|```", "", content).strip()
            items   = json.loads(content)
            return [self._dict_to_result(item) for item in items]
        except json.JSONDecodeError as e:
            logger.error("AI parser returned invalid JSON: %s", e)
            return []
        except Exception as e:
            logger.error("AI requirement parsing failed: %s", e)
            return []

    @staticmethod
    def _dict_to_result(d: dict) -> RequirementResult:
        raw_target = str(d.get("target_id", "s4hana_cloud")).lower().replace(" ", "")
        target_id  = TARGET_ALIASES.get(raw_target, raw_target)
        is_async   = d.get("is_async", False)
        if isinstance(is_async, str):
            is_async = is_async.lower() in ("true", "yes", "1")

        return RequirementResult(
            name=d.get("name", "Unnamed_Interface"),
            sender_system=d.get("sender_system", ""),
            sender_adapter=d.get("sender_adapter", "HTTPS"),
            receiver_system=d.get("receiver_system", ""),
            receiver_adapter=d.get("receiver_adapter", "HTTPS"),
            description=d.get("description", ""),
            namespace=d.get("namespace", ""),
            message_interface=d.get("message_interface", ""),
            mapping_program=d.get("mapping_program", ""),
            target_id=target_id,
            sender_address=d.get("sender_address", ""),
            receiver_address=d.get("receiver_address", ""),
            is_async=is_async,
            message_format=d.get("message_format", "XML"),
            scheduler_cron=d.get("scheduler_cron", ""),
            business_process=d.get("business_process", ""),
            confidence=0.75,
            needs_review=["AI-parsed — review all fields before generating"],
        )

    def parse_docx(self, file_bytes: bytes) -> list[RequirementResult]:
        """Extract text from a .docx file and parse it."""
        try:
            import mammoth
            result = mammoth.extract_raw_text(io.BytesIO(file_bytes))
            return self.parse(result.value)
        except ImportError:
            # Fallback: try as plain text
            try:
                text = file_bytes.decode("utf-8", errors="ignore")
                return self.parse(text)
            except Exception:
                return []


import io


# ---------------------------------------------------------------------------
# Factory
# ---------------------------------------------------------------------------

def parse_requirements(
    source,                         # str path, bytes, or uploaded file content
    source_type: str = "excel",     # "excel" / "text" / "docx"
) -> list[RequirementResult]:
    """
    Unified entry point for the workbench.
    source_type: "excel" | "text" | "docx"
    """
    if source_type == "excel":
        parser = ExcelRequirementParser()
        if isinstance(source, bytes):
            tmp = Path("/tmp/req_upload.xlsx")
            tmp.write_bytes(source)
            return parser.parse(str(tmp))
        return parser.parse(str(source))

    elif source_type == "text":
        parser = AIRequirementParser()
        if isinstance(source, bytes):
            source = source.decode("utf-8", errors="ignore")
        return parser.parse(str(source))

    elif source_type == "docx":
        parser = AIRequirementParser()
        if isinstance(source, str):
            source = Path(source).read_bytes()
        return parser.parse_docx(source)

    else:
        raise ValueError(f"Unknown source_type: {source_type}")


def generate_excel_template() -> bytes:
    """Generate a blank requirements Excel template for download."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl required")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Integration Requirements"

    headers = [
        "Name", "Description", "BusinessProcess",
        "SenderSystem", "SenderAdapter", "SenderAddress",
        "ReceiverSystem", "ReceiverAdapter", "ReceiverAddress",
        "TargetId", "MessageFormat", "Async",
        "MappingProgram", "Namespace", "Scheduler",
        "MessageInterface",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = 22

    # Sample row
    ws.append([
        "PO_to_S4HANA", "Purchase Order replication to S/4HANA Cloud",
        "Procurement",
        "ECC", "IDoc", "http://ecc-host:50000",
        "S4HANA Cloud", "SOAP", "https://tenant.s4hana.cloud.sap",
        "s4hana_cloud", "XML", "false",
        "MM_PO_Create", "http://company.com/po", "",
        "MI_PO_Create",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
