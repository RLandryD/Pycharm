"""
intake/sap_ma_parser.py

Reads SAP Integration Suite Migration Assessment Excel exports and produces
InterfaceRecord lists with adapter/mapping/BPM/multi-mapping flags populated
from real data — no manual column entry required.

The Migration Assessment tool is included free with any Integration Suite
subscription (including free tier). The consultant runs it against the
client's PI/PO system and ships the Excel here. Three-sheet layout:

  Executive Summary    KPI dashboard (Total ICOs, Ready/Adjustment/Eval counts,
                       Total Effort hours). Read for project-level metrics.
  Scenario Evaluation  Per-ICO inventory — main data source. Columns:
                       ICO Technical ID | Sender System | Receiver System |
                       Sender Adapter | Receiver Adapter | Mapping Types Found |
                       Migration Status | Rule Weight | Estimated Effort
  Rules Log            Detected blockers per ICO with remediation strategy.
                       Columns: Triggered Rule ID | Target ICO Scenario |
                       Identified Asset String | Assessment Technical Note.
                       Used to set has_bpm / has_multi_mapping flags when
                       SAP detected those signals automatically.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from extractor.pi_extractor import InterfaceRecord, normalise_adapter

logger = logging.getLogger(__name__)


# Migration Assessment uses SAP technical adapter names; map them to the
# normalised set the rest of the workbench understands.
MA_ADAPTER_ALIASES = {
    "IDOC_AAE":   "IDoc",
    "IDOC":       "IDoc",
    "SOAP":       "SOAP",
    "REST":       "REST",
    "HTTP":       "HTTP",
    "HTTPS":      "HTTPS",
    "FILE":       "File",
    "SFTP":       "SFTP",
    "FTP":        "FTP",
    "JDBC":       "JDBC",
    "JMS":        "JMS",
    "MAIL":       "Mail",
    "RFC":        "RFC",
    "AS2":        "AS2",
    "AS4":        "AS4",
    "HCIC":       "HTTPS",       # HANA Cloud Integration Connector
    "ODATA":      "OData",
    "XI":         "ProcessDirect",
}

# Rules that signal complexity beyond what adapters alone show.
# Migration Assessment emits these in the Rules Log sheet.
COMPLEXITY_SIGNAL_RULES = {
    # Rule ID                       → field, value
    "Java_Mapping_Detected":           ("has_bpm", False),  # Java != BPM, but mark mapping_program
    "BPM_Detected":                    ("has_bpm", True),
    "ccBPM_Detected":                  ("has_bpm", True),
    "Multi_Mapping_Detected":          ("has_multi_mapping", True),
    "Unsupported_Adapter_Module":      ("has_bpm", False),  # implies adapter complexity, not BPM
    "B2B_Seeburger_Legacy":            ("has_multi_mapping", True),
    "Value_Mapping_Context":           (None, None),        # informational
}


@dataclass
class SAPMAExecutiveSummary:
    """KPI dashboard from the Executive Summary sheet."""
    total_icos: int = 0
    ready_to_migrate: int = 0
    adjustment_required: int = 0
    evaluation_required: int = 0
    total_effort_hours: float = 0.0
    raw_rows: list = field(default_factory=list)


@dataclass
class SAPMARule:
    """One triggered rule from the Rules Log sheet."""
    rule_id: str
    affected_ico: str
    asset_string: str
    technical_note: str


@dataclass
class SAPMAReport:
    """Full parsed Migration Assessment output."""
    summary: SAPMAExecutiveSummary
    interfaces: list[InterfaceRecord]
    rules: list[SAPMARule]


def _find_header_row(ws, marker_text: str, max_scan: int = 30,
                     exact: bool = False) -> Optional[int]:
    """Scan first N rows for one containing the marker text in any cell.

    SAP MA exports have title banners on rows 1-3, sometimes a blank row,
    then headers. The exact row varies between versions, so scan for the
    header marker rather than assuming a fixed offset.

    exact=True matches only when a cell's value EQUALS the marker (after
    strip+lower). This avoids matching a long title banner that merely
    contains the marker as a substring (e.g. a row-2 banner
    "Evaluation by Integration Scenario | Source: ..." should NOT be
    mistaken for the real header row whose cell is exactly
    "Integration Scenario").
    """
    marker_lower = marker_text.lower()
    for r in range(1, min(max_scan, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if not val:
                continue
            cell_val = str(val).strip().lower()
            if exact:
                if cell_val == marker_lower:
                    return r
            else:
                if marker_lower in cell_val:
                    return r
    return None



def _parse_effort_hours(value) -> float:
    """Convert '1.5 Hrs', '16.0 Hrs', '420 Hrs', or 1.5 to a float of hours."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    m = re.search(r"([\d.]+)", s)
    return float(m.group(1)) if m else 0.0


def _parse_int(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip().replace(",", "").replace("%", "")
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else 0


def _detect_mapping_signals(mapping_types_str: str) -> tuple[bool, bool, Optional[str]]:
    """Read 'Mapping Types Found' cell and derive flags.

    Returns (has_multi_mapping, mapping_indicator, mapping_program_name)
    where mapping_program_name is a synthetic label since SAP MA reports
    the type, not the original program name.
    """
    if not mapping_types_str:
        return False, False, None
    text = str(mapping_types_str).lower()
    # SAP MA can list multiple mapping types separated by '+' or ','
    has_multi = "+" in text or "," in text or "multi" in text
    has_mapping = bool(text and text != "none" and text != "-")
    # Use the original SAP MA label as the mapping_program name so the
    # downstream complexity analyzer scores it correctly.
    program = str(mapping_types_str).strip() if has_mapping else None
    return has_multi, has_mapping, program


def parse_executive_summary(ws) -> SAPMAExecutiveSummary:
    """Read the KPI dashboard. Header row contains 'Metric' / 'KPI Value'."""
    summary = SAPMAExecutiveSummary()
    header_row = _find_header_row(ws, "Metric")
    if header_row is None:
        logger.warning("Executive Summary: header row not found, skipping")
        return summary

    # Find which column holds 'Metric' and which holds 'KPI Value'
    metric_col = kpi_col = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val is None:
            continue
        v = str(val).strip().lower()
        if v == "metric":
            metric_col = c
        elif "kpi value" in v or v == "value":
            kpi_col = c

    if metric_col is None or kpi_col is None:
        logger.warning("Executive Summary: Metric/KPI columns not found")
        return summary

    metric_map = {
        "total extracted icos":   "total_icos",
        "ready to migrate":       "ready_to_migrate",
        "adjustment required":    "adjustment_required",
        "evaluation required":    "evaluation_required",
        "total estimated effort": "total_effort_hours",
    }
    for r in range(header_row + 1, ws.max_row + 1):
        metric = ws.cell(r, metric_col).value
        kpi    = ws.cell(r, kpi_col).value
        if not metric:
            continue
        key = str(metric).strip().lower()
        field_name = metric_map.get(key)
        if field_name == "total_effort_hours":
            setattr(summary, field_name, _parse_effort_hours(kpi))
        elif field_name:
            setattr(summary, field_name, _parse_int(kpi))
        summary.raw_rows.append((metric, kpi))

    return summary


def parse_rules_log(ws) -> list[SAPMARule]:
    """Read the Rules Log sheet → list of SAPMARule."""
    rules = []
    header_row = _find_header_row(ws, "Triggered Rule")
    if header_row is None:
        logger.warning("Rules Log: header row not found")
        return rules

    # Map headers to column indices. Newer SAP MA layouts insert a
    # 'Complexity' column between ICO Scenario and Asset String — we
    # accept either schema and pull values from whichever columns exist.
    col_idx = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val is None:
            continue
        v = str(val).strip().lower()
        if "triggered rule" in v:
            col_idx["rule_id"] = c
        elif "target ico" in v or "scenario affected" in v:
            col_idx["ico"] = c
        elif v == "complexity":
            col_idx["complexity"] = c
        elif "asset string" in v or "context" in v:
            col_idx["asset"] = c
        elif "technical note" in v or "remediation" in v:
            col_idx["note"] = c

    if "rule_id" not in col_idx:
        return rules

    for r in range(header_row + 1, ws.max_row + 1):
        rid = ws.cell(r, col_idx["rule_id"]).value
        if not rid:
            continue
        rules.append(SAPMARule(
            rule_id=str(rid).strip(),
            affected_ico=str(ws.cell(r, col_idx.get("ico", 0)).value or "").strip(),
            asset_string=str(ws.cell(r, col_idx.get("asset", 0)).value or "").strip(),
            technical_note=str(ws.cell(r, col_idx.get("note", 0)).value or "").strip(),
        ))
    return rules


def parse_scenario_evaluation(ws, rules: list[SAPMARule]) -> list[InterfaceRecord]:
    """Read the Scenario Evaluation sheet → list of InterfaceRecord.

    Each row becomes one record. Rules from the Rules Log sheet are
    cross-referenced to populate has_bpm / has_multi_mapping flags.
    """
    records = []
    header_row = _find_header_row(ws, "ICO Technical ID")
    if header_row is None:
        # Try alternate header
        header_row = _find_header_row(ws, "Sender System")
    if header_row is None:
        logger.warning("Scenario Evaluation: header row not found")
        return records

    # Map columns. SAP MA exports use slightly different column labels
    # between versions — we accept any of the documented variants rather
    # than failing on the first label change. Newer exports add a
    # 'Complexity Group' column (Low/Medium/High Complexity) that
    # SUPERSEDES 'Mapping Types Found'; we read either when present.
    col_idx = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val is None:
            continue
        v = str(val).strip().lower()
        if "ico technical id" in v or v == "ico id":
            col_idx["id"] = c
        elif "sender system" in v:
            col_idx["sender_sys"] = c
        elif "receiver system" in v:
            col_idx["receiver_sys"] = c
        elif "sender adapter" in v:
            col_idx["sender_adapter"] = c
        elif "receiver adapter" in v:
            col_idx["receiver_adapter"] = c
        elif "mapping types" in v or "mapping type" in v:
            col_idx["mapping_types"] = c
        elif "complexity group" in v or v == "complexity":
            # Newer SAP MA layout: 'Complexity Group' replaces 'Mapping Types Found'.
            # We treat it as both the complexity hint AND a mapping-presence signal.
            col_idx["complexity_group"] = c
        elif "migration status" in v:
            col_idx["status"] = c
        elif "rule weight" in v or v == "weight":
            col_idx["weight"] = c
        elif "estimated effort" in v or "est. effort" in v or "est effort" in v or v == "effort":
            col_idx["effort"] = c

    if "id" not in col_idx:
        logger.warning("Scenario Evaluation: ICO ID column not found")
        return records

    # Pre-bucket rules by affected ICO for fast lookup
    rules_by_ico: dict[str, list[SAPMARule]] = {}
    for rule in rules:
        rules_by_ico.setdefault(rule.affected_ico, []).append(rule)

    def cell(r, key):
        c = col_idx.get(key)
        return ws.cell(r, c).value if c else None

    for r in range(header_row + 1, ws.max_row + 1):
        ico_id = cell(r, "id")
        if not ico_id:
            continue
        ico_id_str = str(ico_id).strip()

        sender_adapter_raw   = str(cell(r, "sender_adapter") or "").strip().upper()
        receiver_adapter_raw = str(cell(r, "receiver_adapter") or "").strip().upper()
        sender_adapter   = MA_ADAPTER_ALIASES.get(sender_adapter_raw,
                                                   normalise_adapter(sender_adapter_raw or "HTTPS"))
        receiver_adapter = MA_ADAPTER_ALIASES.get(receiver_adapter_raw,
                                                   normalise_adapter(receiver_adapter_raw or "HTTPS"))

        mapping_types = cell(r, "mapping_types")
        complexity_group_raw = cell(r, "complexity_group")
        if mapping_types:
            has_multi, has_mapping, mapping_program = _detect_mapping_signals(mapping_types)
        elif complexity_group_raw:
            # Newer SAP MA layout: 'Complexity Group' carries Low/Medium/High
            # Complexity. Treat any Medium/High as having a mapping (since
            # SAP would not classify above Low without transformation logic
            # being present), but don't infer multi-mapping from it alone.
            cg = str(complexity_group_raw).strip().lower()
            has_mapping = "medium" in cg or "high" in cg
            has_multi   = False
            mapping_program = (f"SAP MA: {complexity_group_raw}"
                               if has_mapping else None)
        else:
            has_multi, has_mapping, mapping_program = False, False, None

        # Pull a Low/Medium/High complexity hint when SAP provided one.
        # We don't overwrite our analyzer's score with it — that gets
        # computed separately — but we surface it in description for
        # traceability and as the raw['sap_ma_complexity'] field.
        sap_complexity = ""
        if complexity_group_raw:
            cg = str(complexity_group_raw).strip().lower()
            if "high" in cg:
                sap_complexity = "HIGH"
            elif "medium" in cg:
                sap_complexity = "MEDIUM"
            elif "low" in cg:
                sap_complexity = "LOW"

        # Cross-reference rules log for complexity signals
        ico_rules = rules_by_ico.get(ico_id_str, [])
        has_bpm = False
        for rule in ico_rules:
            field_name, flag_value = COMPLEXITY_SIGNAL_RULES.get(
                rule.rule_id, (None, None))
            if field_name == "has_bpm" and flag_value:
                has_bpm = True
            elif field_name == "has_multi_mapping" and flag_value:
                has_multi = True

        # SAP MA's estimated effort is in hours; the analyzer expects days
        # for its own metric but we keep this in description for traceability.
        effort_hrs = _parse_effort_hours(cell(r, "effort"))
        status     = str(cell(r, "status") or "").strip()
        weight     = _parse_int(cell(r, "weight"))

        # Channel count: SAP MA doesn't report this directly, but Rule Weight
        # is a reasonable proxy (higher weight = more configuration). Bucket:
        #   weight < 30  → 1 channel
        #   weight < 80  → 2 channels
        #   weight >= 80 → 3 channels
        if weight < 30:
            channel_count = 1
        elif weight < 80:
            channel_count = 2
        else:
            channel_count = 3

        description_parts = [f"SAP MA status: {status}"] if status else []
        if sap_complexity:
            description_parts.append(f"SAP complexity: {sap_complexity}")
        if effort_hrs:
            description_parts.append(f"SAP estimated effort: {effort_hrs} hrs")
        if ico_rules:
            description_parts.append(f"{len(ico_rules)} rule(s) triggered")
        description = " | ".join(description_parts)

        records.append(InterfaceRecord(
            id=re.sub(r"[^\w]", "_", ico_id_str)[:60],
            name=ico_id_str[:80],
            namespace="",
            software_component="",
            sender_system=str(cell(r, "sender_sys") or "").strip(),
            receiver_system=str(cell(r, "receiver_sys") or "").strip(),
            sender_adapter=sender_adapter,
            receiver_adapter=receiver_adapter,
            message_interface=ico_id_str,
            mapping_program=mapping_program,
            description=description,
            has_bpm=has_bpm,
            has_multi_mapping=has_multi,
            channel_count=channel_count,
            ma_weight=weight if weight else None,
            ma_size=__import__("analyzer.sap_complexity_engine",
                               fromlist=["ma_size_from_group_status"]
                               ).ma_size_from_group_status(
                str(complexity_group_raw or ""), status),
            ma_status=status,
            raw={
                "sap_ma_status":     status,
                "sap_ma_weight":     weight,
                "sap_ma_effort_hrs": effort_hrs,
                "sap_ma_complexity": sap_complexity,
                "sap_ma_rules":      [r.rule_id for r in ico_rules],
                "sap_ma_rule_assets": [[r.rule_id, r.asset_string]
                                       for r in ico_rules],
            },
        ))

    return records


def _parse_realschema_scenarios(ws, rules: list[SAPMARule]) -> tuple[list[InterfaceRecord], SAPMAExecutiveSummary]:
    """Parse the REAL SAP MA schema 'Evaluation by Scenario' sheet.

    This layout (matches the actual SAP export + our mock generator) uses:
      Integration Scenario | Sender System | Receiver System |
      Sender Adapter | Receiver Adapter | Weight | T-Shirt Size |
      Assessment Category | Min/Avg/Max Effort (Hrs)

    Assessment Category values: 'Ready to migrate' / 'Adjustments required' /
    'Evaluation required'. Effort is per-interface in hours. Returns the
    records AND a computed executive summary (this schema has no separate
    summary sheet — the KPIs are derived from the rows).
    """
    records: list[InterfaceRecord] = []
    summary = SAPMAExecutiveSummary()

    header_row = _find_header_row(ws, "Integration Scenario", exact=True)
    if header_row is None:
        logger.warning("Real-schema: 'Integration Scenario' header not found")
        return records, summary

    col_idx = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val is None:
            continue
        v = str(val).strip().lower()
        if "integration scenario" in v:
            col_idx["id"] = c
        elif "sender system" in v:
            col_idx["sender_sys"] = c
        elif "receiver system" in v:
            col_idx["receiver_sys"] = c
        elif "sender adapter" in v:
            col_idx["sender_adapter"] = c
        elif "receiver adapter" in v:
            col_idx["receiver_adapter"] = c
        elif v == "weight":
            col_idx["weight"] = c
        elif "t-shirt" in v or "tshirt" in v or "t shirt" in v:
            col_idx["tshirt"] = c
        elif "assessment category" in v or v == "category":
            col_idx["category"] = c
        elif "min effort" in v:
            col_idx["min_effort"] = c
        elif "avg effort" in v or "average effort" in v:
            col_idx["avg_effort"] = c
        elif "max effort" in v:
            col_idx["max_effort"] = c

    if "id" not in col_idx:
        logger.warning("Real-schema: Integration Scenario column not found")
        return records, summary

    rules_by_ico: dict[str, list[SAPMARule]] = {}
    for rule in rules:
        rules_by_ico.setdefault(rule.affected_ico, []).append(rule)

    def cell(r, key):
        c = col_idx.get(key)
        return ws.cell(r, c).value if c else None

    ready = adjustment = evaluation = 0
    total_effort = 0.0

    for r in range(header_row + 1, ws.max_row + 1):
        ico = cell(r, "id")
        if not ico:
            continue
        ico_str = str(ico).strip()
        # Skip any totals/footer row
        if ico_str.lower().startswith(("project total", "total")):
            continue

        category = str(cell(r, "category") or "").strip()
        cat_low = category.lower()
        # Map category -> our complexity + KPI buckets
        if "ready" in cat_low:
            ready += 1
            complexity_hint = "LOW"
        elif "adjustment" in cat_low:
            adjustment += 1
            complexity_hint = "MEDIUM"
        elif "evaluation" in cat_low:
            evaluation += 1
            complexity_hint = "HIGH"
        else:
            complexity_hint = ""

        # Effort: prefer avg, fall back to max then min
        avg_e = _parse_effort_hours(cell(r, "avg_effort"))
        eff = avg_e or _parse_effort_hours(cell(r, "max_effort")) or _parse_effort_hours(cell(r, "min_effort"))
        total_effort += eff

        sender_adapter_raw   = str(cell(r, "sender_adapter") or "").strip().upper()
        receiver_adapter_raw = str(cell(r, "receiver_adapter") or "").strip().upper()
        sender_adapter   = MA_ADAPTER_ALIASES.get(sender_adapter_raw,
                                                   normalise_adapter(sender_adapter_raw or "HTTPS"))
        receiver_adapter = MA_ADAPTER_ALIASES.get(receiver_adapter_raw,
                                                   normalise_adapter(receiver_adapter_raw or "HTTPS"))

        ico_rules = rules_by_ico.get(ico_str, [])
        has_bpm = any("bpm" in (rule.rule_id or "").lower() or
                      "ccbpm" in (rule.technical_note or "").lower()
                      for rule in ico_rules)

        tshirt = str(cell(r, "tshirt") or "").strip()
        weight = _parse_int(cell(r, "weight"))

        # Medium/High assessment implies transformation logic present
        has_mapping = complexity_hint in ("MEDIUM", "HIGH")
        mapping_program = (f"SAP MA: {category}" if has_mapping else None)

        desc_parts = []
        if category:        desc_parts.append(f"SAP category: {category}")
        if complexity_hint: desc_parts.append(f"SAP complexity: {complexity_hint}")
        if tshirt:          desc_parts.append(f"T-shirt: {tshirt}")
        if eff:             desc_parts.append(f"SAP effort: {eff} hrs")
        if ico_rules:       desc_parts.append(f"{len(ico_rules)} rule(s)")

        records.append(InterfaceRecord(
            id=re.sub(r"[^\w]", "_", ico_str)[:60],
            name=ico_str[:80],
            namespace="",
            software_component="",
            sender_system=str(cell(r, "sender_sys") or "").strip(),
            receiver_system=str(cell(r, "receiver_sys") or "").strip(),
            sender_adapter=sender_adapter,
            receiver_adapter=receiver_adapter,
            message_interface=ico_str,
            mapping_program=mapping_program,
            description=" | ".join(desc_parts),
            has_bpm=has_bpm,
            has_multi_mapping=False,
            channel_count=1,
            ma_weight=weight if weight else None,
            ma_size=tshirt,
            ma_status=category,
            raw={
                "sap_ma_category":   category,
                "sap_ma_complexity": complexity_hint,
                "sap_ma_tshirt":     tshirt,
                "sap_ma_weight":     weight,
                "sap_ma_effort_hrs": eff,
                "sap_ma_rules":      [rr.rule_id for rr in ico_rules],
                "sap_ma_rule_assets": [[rr.rule_id, rr.asset_string]
                                       for rr in ico_rules],
            },
        ))

    summary.total_icos          = len(records)
    summary.ready_to_migrate    = ready
    summary.adjustment_required = adjustment
    summary.evaluation_required = evaluation
    summary.total_effort_hours  = total_effort
    return records, summary


def _parse_realschema_rules(ws) -> list[SAPMARule]:
    """Parse the REAL SAP MA 'Full Evaluation Results' sheet → rules.

    Columns: Integration Scenario | Triggered Rule | Rule Category |
    Rule Weight | Interfaces Engaged | Remediation Note.
    """
    rules: list[SAPMARule] = []
    header_row = _find_header_row(ws, "Triggered Rule", exact=True)
    if header_row is None:
        return rules
    col_idx = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val is None:
            continue
        v = str(val).strip().lower()
        if "integration scenario" in v:
            col_idx["ico"] = c
        elif "triggered rule" in v:
            col_idx["rule_id"] = c
        elif "interfaces engaged" in v or "asset" in v or "context" in v:
            col_idx["asset"] = c
        elif "remediation" in v or "technical note" in v:
            col_idx["note"] = c
    if "rule_id" not in col_idx or "ico" not in col_idx:
        return rules
    def cell(r, key):
        c = col_idx.get(key)
        return ws.cell(r, c).value if c else None
    for r in range(header_row + 1, ws.max_row + 1):
        rid = cell(r, "rule_id")
        ico = cell(r, "ico")
        if not rid or not ico:
            continue
        rules.append(SAPMARule(
            rule_id=str(rid).strip(),
            affected_ico=str(ico).strip(),
            asset_string=str(cell(r, "asset") or "").strip(),
            technical_note=str(cell(r, "note") or "").strip(),
        ))
    return rules


def parse_sap_ma_excel(path: str | Path) -> SAPMAReport:
    """Parse a SAP Migration Assessment Excel export end-to-end.

    Returns a SAPMAReport bundling the executive summary, per-ICO records,
    and the rules log. Raises ValueError if the file doesn't match the
    expected three-sheet structure.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel parsing")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # Locate sheets by name (case-insensitive, partial match for resilience
    # since SAP could rename them between versions)
    sheet_map = {s.lower(): s for s in wb.sheetnames}
    def find_sheet(needle: str) -> Optional[str]:
        for low, real in sheet_map.items():
            if needle.lower() in low:
                return real
        return None

    # ── Schema detection ─────────────────────────────────────────────────
    # Two schemas supported:
    #   REAL    : 'Evaluation by Scenario' + 'Full Evaluation Results'
    #             (matches the actual SAP export + our mock generator)
    #   LEGACY  : 'Scenario Evaluation' + 'Rules Log' + 'Executive Summary'
    #             (the older simulation files)
    # We detect by sheet presence and route accordingly. Content-signature
    # based so a renamed tab still works (we also check column markers).
    real_scenario = (find_sheet("evaluation by scenario")
                     or find_sheet("evaluation by integration"))
    real_results  = (find_sheet("full evaluation results")
                     or find_sheet("full evaluation"))

    # If the obvious real-schema tabs aren't named as expected, sniff any
    # sheet for the 'Integration Scenario' + 'Assessment Category' markers.
    if real_scenario is None:
        for s in wb.sheetnames:
            ws = wb[s]
            if (_find_header_row(ws, "Integration Scenario", exact=True) is not None
                    and _find_header_row(ws, "Assessment Category", exact=True) is not None):
                real_scenario = s
                break

    if real_scenario is not None:
        # REAL schema path
        rules = (_parse_realschema_rules(wb[real_results])
                 if real_results else [])
        records, summary = _parse_realschema_scenarios(wb[real_scenario], rules)
        logger.info("Parsed SAP MA (real schema): %d interfaces, %d rules",
                    len(records), len(rules))
        return SAPMAReport(summary=summary, interfaces=records, rules=rules)

    # ── LEGACY schema path ───────────────────────────────────────────────
    summary_name  = find_sheet("executive summary")
    scenario_name = find_sheet("scenario evaluation")
    rules_name    = find_sheet("rules log")

    if scenario_name is None:
        raise ValueError(
            "File does not look like a SAP Migration Assessment export. "
            "Expected either the real schema ('Evaluation by Scenario' + "
            "'Full Evaluation Results') or the legacy schema "
            "('Scenario Evaluation' + 'Rules Log'). "
            f"Sheets present: {wb.sheetnames}")

    summary = (parse_executive_summary(wb[summary_name])
               if summary_name else SAPMAExecutiveSummary())
    rules   = parse_rules_log(wb[rules_name]) if rules_name else []
    records = parse_scenario_evaluation(wb[scenario_name], rules)

    logger.info("Parsed SAP MA export (legacy schema): %d interfaces, %d rules, %d ICOs total",
                len(records), len(rules), summary.total_icos)

    return SAPMAReport(summary=summary, interfaces=records, rules=rules)


