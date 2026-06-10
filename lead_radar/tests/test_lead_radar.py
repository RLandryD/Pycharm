import json
import pathlib
from datetime import date

import pytest
from openpyxl import load_workbook

from lead_radar.exporter import export_xlsx
from lead_radar.models import Posting
from lead_radar.people import (build_search_links, match_role,
                               parse_cse_payload, xray_query)
from lead_radar.scoring import (aggregate_companies, classify_company, norm,
                                score_posting)
from lead_radar.sources.adzuna import parse_adzuna_payload
from lead_radar.sources.jsearch import parse_jsearch_payload

FIX = pathlib.Path(__file__).parent / "fixtures"
TODAY = date(2026, 6, 9)


def load(name):
    return json.loads((FIX / name).read_text(encoding="utf-8"))


# ---- scoring -----------------------------------------------------------------

def test_norm_strips_accents():
    assert norm("Migración México") == "migracion mexico"


def test_pipo_migration_posting_scores_high():
    p = Posting(source="t", company="X", title="Líder SAP CPI",
                description="Migración de SAP PI/PO hacia Integration Suite. iFlows Groovy IDoc.",
                location="Estado de México", posted_date="2026-05-20")
    score_posting(p, today=TODAY)
    # 40 pipo + 20 cpi + 25 migration + depth + 10 recency + 5 mx
    assert p.score >= 100
    assert any("PI/PO" in r for r in p.score_reasons)


def test_irrelevant_posting_scores_zero():
    p = Posting(source="t", company="Acme", title="Panadero",
                description="Hacer pan dulce", posted_date="2026-06-01")
    score_posting(p, today=TODAY)
    assert p.score <= 15  # at most recency, no stack signals


def test_cpi_only_scores_medium():
    p = Posting(source="t", company="X", title="SAP CPI Developer",
                description="Integration Suite iFlows", posted_date="2025-01-01")
    score_posting(p, today=TODAY)
    assert 20 <= p.score < 40


def test_recency_bonus_tiers():
    recent = score_posting(Posting("t", "A", "SAP CPI dev", description="sap cpi",
                                   posted_date="2026-06-01"), today=TODAY)
    old = score_posting(Posting("t", "B", "SAP CPI dev", description="sap cpi",
                                posted_date="2024-01-01"), today=TODAY)
    assert recent.score == old.score + 10


# ---- classification / aggregation ---------------------------------------------

def test_channel_classification():
    assert classify_company("Infosys") == "channel"
    assert classify_company("EAS Consulting SA de CV") == "channel"
    assert classify_company("Grupo Truper") == "end_client"


def test_aggregate_merges_company_suffix_variants():
    p1 = score_posting(Posting("t", "Grupo Truper", "Líder SAP CPI",
                               description="pi/po migración integration suite",
                               posted_date="2026-05-20"), today=TODAY)
    p2 = score_posting(Posting("t", "Grupo Truper SA de CV", "CPI Developer",
                               description="sap cpi iflow", posted_date="2026-05-28"),
                       today=TODAY)
    leads = aggregate_companies([p1, p2])
    assert len(leads) == 1
    assert leads[0].n_postings == 2
    assert leads[0].score == p1.score + 5  # best + multi-posting bonus


def test_aggregate_ranks_descending():
    strong = score_posting(Posting("t", "A", "PI/PO migration to Integration Suite",
                                   description="pi/po migration integration suite groovy idoc",
                                   posted_date="2026-06-01"), today=TODAY)
    weak = score_posting(Posting("t", "B", "SAP CPI", description="sap cpi",
                                 posted_date="2024-01-01"), today=TODAY)
    leads = aggregate_companies([weak, strong])
    assert [l.company for l in leads] == ["A", "B"]


# ---- source parsers ------------------------------------------------------------

def test_parse_jsearch_fixture():
    postings = parse_jsearch_payload(load("jsearch_sample.json"), "q")
    assert len(postings) == 4
    truper = postings[0]
    assert truper.company == "Grupo Truper"
    assert truper.posted_date == "2026-05-20"
    assert "Jilotepec" in truper.location


def test_parse_adzuna_fixture():
    postings = parse_adzuna_payload(load("adzuna_sample.json"), "q")
    assert len(postings) == 2
    assert postings[1].company == "KION Group"
    assert postings[1].posted_date == "2025-06-01"


def test_end_to_end_fixture_ranking():
    postings = (parse_jsearch_payload(load("jsearch_sample.json"), "q")
                + parse_adzuna_payload(load("adzuna_sample.json"), "q"))
    for p in postings:
        score_posting(p, today=TODAY)
    leads = aggregate_companies(postings)
    names = [l.company for l in leads]
    assert names[0].startswith("Grupo Truper")          # strongest signal
    assert names.index("Grupo Truper") < names.index("Acme Bakery")
    infosys = next(l for l in leads if l.company == "Infosys")
    assert infosys.lead_type == "channel"


# ---- people --------------------------------------------------------------------

def test_xray_query_shape():
    q = xray_query("CEMEX", ["sap delivery manager", "gerente de entrega sap"])
    assert q.startswith('site:linkedin.com/in "CEMEX"')
    assert '"sap delivery manager"' in q


def test_match_role_priority():
    rank, label = match_role("Juan - SAP Integration Architect - CEMEX")
    assert rank == 4 and "Architect" in label
    rank, _ = match_role("SAP Delivery Manager LATAM")
    assert rank == 1
    rank, _ = match_role("Marketing Coordinator")
    assert rank is None


def test_parse_cse_extracts_people():
    people = parse_cse_payload(load("cse_sample.json"), "CEMEX")
    assert len(people) == 3
    arch = people[0]
    assert arch.name == "Juan Pérez"
    assert arch.role_rank == 4
    dm = people[1]
    assert dm.role_rank == 1
    assert people[2].role_rank is None  # marketing -> manual filter


def test_build_search_links_count():
    links = build_search_links(["CEMEX", "Truper"], top_roles=6)
    assert len(links) == 12
    assert links[0][1] == 1  # rank ordering preserved


# ---- exporter ------------------------------------------------------------------

def test_export_xlsx_smoke(tmp_path):
    postings = parse_jsearch_payload(load("jsearch_sample.json"), "q")
    for p in postings:
        score_posting(p, today=TODAY)
    leads = aggregate_companies(postings)
    people = parse_cse_payload(load("cse_sample.json"), "CEMEX")
    links = build_search_links([l.company for l in leads], top_roles=3)
    out = tmp_path / "leads.xlsx"
    export_xlsx(str(out), leads, people, links)

    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Companies", "Postings", "People", "SearchLinks"}
    comp = wb["Companies"]
    assert comp.max_row == len(leads) + 1
    assert comp.cell(row=1, column=1).value == "Score"
    ppl = wb["People"]
    assert ppl.max_row == len(people) + 1


# ---- setup wizard ---------------------------------------------------------------

from lead_radar.setup_wizard import interpret_status, mask, write_env, run_setup


def test_interpret_status_mapping():
    assert interpret_status("jsearch", 200) == (True, "valid")
    ok, detail = interpret_status("jsearch", 403)
    assert ok is False and "403" in detail
    ok, _ = interpret_status("google_cse", 429)
    assert ok is True  # key works, just throttled
    ok, _ = interpret_status("adzuna", 404)
    assert ok is False
    ok, _ = interpret_status("jsearch", 500)
    assert ok is None  # inconclusive, save-but-warn


def test_mask_never_reveals_full_key():
    assert mask("abcdefgh12345") == "****2345"
    assert "abcdefgh" not in mask("abcdefgh12345")


def test_write_env_merges_and_preserves(tmp_path):
    env = tmp_path / ".env"
    env.write_text("RAPIDAPI_KEY=old\n# comment kept\nOTHER=x\n")
    write_env({"RAPIDAPI_KEY": "newkey", "GOOGLE_CSE_KEY": "g1"}, path=str(env))
    text = env.read_text()
    assert "RAPIDAPI_KEY=newkey" in text
    assert "# comment kept" in text
    assert "OTHER=x" in text
    assert "GOOGLE_CSE_KEY=g1" in text
    mode = env.stat().st_mode & 0o777
    assert mode == 0o600


def test_setup_wizard_skip_all(tmp_path, monkeypatch, capsys):
    monkeypatch.chdir(tmp_path)
    rc = run_setup(open_browser=False, input_fn=lambda prompt: "")
    assert rc == 0
    out = capsys.readouterr().out
    assert "Nothing saved" in out
    assert not (tmp_path / ".env").exists()


def test_setup_wizard_rejects_invalid_key(tmp_path, monkeypatch, capsys):
    monkeypatch.chdir(tmp_path)
    answers = iter(["badkey", "", "", ""])  # jsearch key, skip serper/google/adzuna
    import lead_radar.setup_wizard as sw
    monkeypatch.setattr(sw, "validate_jsearch",
                        lambda key, timeout=20: (False, "rejected by jsearch (HTTP 403)"))
    rc = run_setup(open_browser=False, input_fn=lambda p: next(answers))
    assert rc == 0
    assert "INVALID" in capsys.readouterr().out
    assert not (tmp_path / ".env").exists()


def test_setup_wizard_saves_valid_key(tmp_path, monkeypatch, capsys):
    monkeypatch.chdir(tmp_path)
    answers = iter(["goodkey123", "", "", ""])
    import lead_radar.setup_wizard as sw
    monkeypatch.setattr(sw, "validate_jsearch",
                        lambda key, timeout=20: (True, "valid"))
    rc = run_setup(open_browser=False, input_fn=lambda p: next(answers))
    assert rc == 0
    env = (tmp_path / ".env").read_text()
    assert "RAPIDAPI_KEY=goodkey123" in env
    out = capsys.readouterr().out
    assert "OK (verified)" in out


# ---- serper provider ------------------------------------------------------------

from lead_radar.people import parse_serper_payload, people_provider
from lead_radar.config import Settings


def test_parse_serper_payload():
    payload = {"organic": [
        {"title": "Juan Pérez - SAP Integration Architect - CEMEX | LinkedIn",
         "link": "https://mx.linkedin.com/in/juanperez",
         "snippet": "SAP CPI, PI/PO. Monterrey."},
        {"title": "Ana Ruiz - SAP Delivery Manager - CEMEX | LinkedIn",
         "link": "https://mx.linkedin.com/in/anaruiz",
         "snippet": "Delivery manager."},
    ]}
    people = parse_serper_payload(payload, "CEMEX")
    assert len(people) == 2
    assert people[0].name == "Juan Pérez" and people[0].role_rank == 4
    assert people[1].role_rank == 1


def test_people_provider_priority(monkeypatch):
    s = Settings()
    s.serper_api_key, s.google_cse_key, s.google_cse_cx = "x", "y", "z"
    assert people_provider(s) == "serper"
    s.serper_api_key = ""
    assert people_provider(s) == "google_cse"
    s.google_cse_key = ""
    assert people_provider(s) == ""


# ---- blocklist & recruiter filtering --------------------------------------------

from lead_radar.scoring import load_blocklist, is_blocked
from lead_radar.people import is_recruiter, XRAY_EXCLUSIONS
from lead_radar.models import PersonCandidate


def test_blocklist_load_and_match(tmp_path):
    bl = tmp_path / "blocklist.txt"
    bl.write_text("# comment\nXideral\nDeloitte\nAllianceIT\n")
    blocked = load_blocklist(str(bl))
    assert is_blocked("Xideral SA de CV", blocked)
    assert is_blocked("Deloitte Consulting México", blocked)
    assert is_blocked("ALLIANCEIT", blocked)
    assert not is_blocked("Grupo Truper", blocked)


def test_blocklist_missing_file_is_empty():
    assert load_blocklist("/nonexistent/blocklist.txt") == set()


def test_channel_detection_new_firms():
    from lead_radar.scoring import classify_company
    assert classify_company("Xideral") == "channel"
    assert classify_company("AllianceIT") == "channel"


def test_recruiter_filter():
    rec = PersonCandidate(company="X", name="A", title="Senior IT Recruiter",
                          snippet="Talent acquisition for SAP roles")
    arch = PersonCandidate(company="X", name="B", title="SAP Integration Architect",
                           snippet="PI/PO and CPI")
    hr_es = PersonCandidate(company="X", name="C", title="Atracción de Talento TI",
                            snippet="")
    assert is_recruiter(rec)
    assert is_recruiter(hr_es)
    assert not is_recruiter(arch)


def test_xray_query_includes_exclusions():
    from lead_radar.people import xray_query
    q = xray_query("CEMEX", ["sap delivery manager"])
    assert "-recruiter" in q and '-"talent acquisition"' in q
